# Imports estándar de Python
import os
import io
import shutil
import secrets
from datetime import datetime, date, timedelta
import calendar  # ✅ NUEVO: Para calcular último día del mes
from typing import Optional
# Imports de librerías de terceros (pypi)
from fastapi import FastAPI, Depends, HTTPException, Request, Form, Query, UploadFile, File
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session
import pandas as pd

# Imports de módulos locales de la aplicación
import models
import database
import auth
import excel_import
from models import TipoUsuario, EstadoProspecto
from sqlalchemy import func, or_, and_
from difflib import get_close_matches

# Imports de utilidades locales
from utils import (
    parsear_fecha,
    calcular_rango_fechas,
    normalizar_fecha_input,
    normalizar_texto_mayusculas,
    normalizar_numero,
    normalizar_email,
    enviar_notificacion_email
)


app = FastAPI(title="Sistema de Prospectos")

# Configuración para upload de archivos
UPLOAD_DIR = "uploads"
os.makedirs(UPLOAD_DIR, exist_ok=True)

app.mount("/static", StaticFiles(directory="static"), name="static")
app.mount("/uploads", StaticFiles(directory=UPLOAD_DIR), name="uploads")
templates = Jinja2Templates(directory="templates")

# Almacenamiento simple de sesiones en memoria
active_sessions = {}

# Crear tablas al inicio
@app.on_event("startup")
def startup():
    database.create_tables()
    db = next(database.get_db())
    try:
        # Crear medios de ingreso por defecto
        medios = ["REDES", "TEL TRAVEL", "RECOMPRA", "REFERIDO", "FIDELIZACION"]
        for medio in medios:
            if not db.query(models.MedioIngreso).filter(models.MedioIngreso.nombre == medio).first():
                db.add(models.MedioIngreso(nombre=medio))
        
        # Crear usuario administrador por defecto
        if not db.query(models.Usuario).filter(models.Usuario.username == "admin").first():
            admin_user = models.Usuario(
                username="admin",
                email="admin@empresa.com",
                hashed_password=auth.get_password_hash("admin123"),
                tipo_usuario=TipoUsuario.ADMINISTRADOR.value
            )
            db.add(admin_user)
        
        # Crear un agente de prueba
        if not db.query(models.Usuario).filter(models.Usuario.username == "agente1").first():
            agente_user = models.Usuario(
                username="agente1",
                email="agente1@empresa.com",
                hashed_password=auth.get_password_hash("agente123"),
                tipo_usuario=TipoUsuario.AGENTE.value
            )
            db.add(agente_user)
        
        # Crear usuario de Servicio al Cliente
        if not db.query(models.Usuario).filter(models.Usuario.username == "servicio_cliente").first():
            servicio_user = models.Usuario(
                username="servicio_cliente",
                email="servicioclientetravelhouse@gmail.com",
                hashed_password=auth.get_password_hash("servicio123"),
                tipo_usuario=TipoUsuario.AGENTE.value,
                activo=1
            )
            db.add(servicio_user)
        
        db.commit()
        print("Datos iniciales creados correctamente")
        print("Usuario admin: admin / admin123")
        print("Usuario agente: agente1 / agente123")
        print("Usuario servicio_cliente: servicio_cliente / servicio123")
        print("No se crearon prospectos de prueba. Puedes crearlos manualmente.")
        
    except Exception as e:
        db.rollback()
        print(f"❌ Error inicializando datos: {e}")
    finally:
        db.close()




# Función simple para obtener usuario actual
async def get_current_user(request: Request, db: Session = Depends(database.get_db)):
    try:
        session_token = request.cookies.get("session_token")
        
        if not session_token:
            return None
        
        user_id = active_sessions.get(session_token)
        if not user_id:
            return None
        
        user = db.query(models.Usuario).filter(models.Usuario.id == user_id).first()
        return user
    except Exception as e:
        print(f"❌ Error in get_current_user: {e}")
        return None

# Verificar si usuario es admin
async def require_admin(user: models.Usuario = Depends(get_current_user)):
    if not user or user.tipo_usuario != TipoUsuario.ADMINISTRADOR.value:
        raise HTTPException(status_code=403, detail="No tiene permisos de administrador")
    return user

# ========== FUNCIÓN PARA NOTIFICACIONES AUTOMÁTICAS DE VIAJE ==========

def crear_notificaciones_viaje(prospecto: models.Prospecto, db: Session):
    """
    Crea notificaciones automáticas para seguimiento de viaje.
    
    Genera 3 notificaciones:
    - 45 días antes: Confirmar pagos y estado de reserva
    - 10 días antes: Validar gestiones pre-viaje
    - 2 días antes: Validar pre-viaje, formularios y gestiones finales
    
    Args:
        prospecto: Objeto Prospecto con estado ganado
        db: Sesión de base de datos
    """
    # Validar que tenga fecha_ida
    if not prospecto.fecha_ida:
        return
    
    # Validar que sea estado ganado
    if prospecto.estado != EstadoProspecto.GANADO.value:
        return
    
    hoy = datetime.now().date()
    
    # Eliminar notificaciones automáticas anteriores de este prospecto
    db.query(models.Notificacion).filter(
        models.Notificacion.prospecto_id == prospecto.id,
        models.Notificacion.tipo == 'seguimiento_viaje'
    ).delete()
    
    # Definir notificaciones
    notificaciones_config = [
        {
            'dias_antes': 45,
            'mensaje': f'Confirmar pagos y estado de reserva - Viaje a {prospecto.destino or "destino"}'
        },
        {
            'dias_antes': 10,
            'mensaje': f'Validar con cliente gestiones pre-viaje - Viaje a {prospecto.destino or "destino"}'
        },
        {
            'dias_antes': 2,
            'mensaje': f'Validar pre-viaje, formularios y gestiones finales - Viaje a {prospecto.destino or "destino"}'
        }
    ]
    
    # Crear notificaciones
    for config in notificaciones_config:
        fecha_notificacion = prospecto.fecha_ida - timedelta(days=config['dias_antes'])
        
        # Solo crear si la fecha es futura
        if fecha_notificacion >= hoy:
            nueva_notificacion = models.Notificacion(
                prospecto_id=prospecto.id,
                usuario_id=prospecto.agente_asignado_id,
                tipo='seguimiento_viaje',
                mensaje=config['mensaje'],
                fecha_programada=datetime.combine(fecha_notificacion, datetime.min.time()),
                leida=False,
                email_enviado=False
            )
            db.add(nueva_notificacion)
    
    db.commit()

# ========== FUNCIÓN PARA ORGANIZAR UPLOADS POR FECHA ==========

def obtener_ruta_upload_por_fecha(fecha: datetime = None) -> str:
    """
    Genera la ruta de directorio para uploads basada en la fecha.
    
    Estructura: uploads/YYYY/MM/DD/
    
    Args:
        fecha: Fecha para generar la ruta. Si es None, usa la fecha actual.
    
    Returns:
        str: Ruta del directorio (ej: "uploads/2026/01/10/")
    """
    if fecha is None:
        fecha = datetime.now()
    
    # Crear estructura: uploads/YYYY/MM/DD/
    ruta = os.path.join(
        UPLOAD_DIR,
        str(fecha.year),
        f"{fecha.month:02d}",
        f"{fecha.day:02d}"
    )
    
    # Crear directorios si no existen
    os.makedirs(ruta, exist_ok=True)
    
    return ruta

# Página de login
@app.get("/", response_class=HTMLResponse)
async def login_page(request: Request):
    return templates.TemplateResponse("login.html", {"request": request})

@app.post("/login")
async def login(
    request: Request, 
    username: str = Form(...), 
    password: str = Form(...), 
    db: Session = Depends(database.get_db)
):
    user = db.query(models.Usuario).filter(models.Usuario.username == username).first()
    if not user:
        return templates.TemplateResponse("login.html", {
            "request": request, 
            "error": "Usuario no encontrado"
        })
    
    if not auth.verify_password(password, user.hashed_password):
        return templates.TemplateResponse("login.html", {
            "request": request, 
            "error": "Contraseña incorrecta"
        })
    
    session_token = secrets.token_urlsafe(32)
    active_sessions[session_token] = user.id
    
    response = RedirectResponse(url="/dashboard", status_code=303)
    response.set_cookie(
        key="session_token",
        value=session_token,
        httponly=True,
        max_age=1800,
        path="/"
    )
    
    return response

# Logout
@app.get("/logout")
async def logout(request: Request):
    session_token = request.cookies.get("session_token")
    if session_token and session_token in active_sessions:
        del active_sessions[session_token]
    
    response = RedirectResponse(url="/", status_code=303)
    response.delete_cookie("session_token")
    return response


# ========== IMPORTACIÓN DE DATOS DESDE EXCEL ==========

@app.get("/importar-datos", response_class=HTMLResponse)
async def mostrar_importar_datos(
    request: Request,
    db: Session = Depends(database.get_db),
    user: models.Usuario = Depends(require_admin)
):
    """Muestra la página de importación de datos (solo administradores)"""
    return templates.TemplateResponse("importar_datos.html", {
        "request": request,
        "current_user": user
    })


@app.post("/importar-usuarios")
async def importar_usuarios(
    request: Request,
    archivo: UploadFile = File(...),
    db: Session = Depends(database.get_db),
    user: models.Usuario = Depends(require_admin)
):
    """Procesa la importación de usuarios desde Excel"""
    try:
        # Validar archivo
        es_valido, mensaje_error = excel_import.validar_archivo_excel(archivo)
        if not es_valido:
            return templates.TemplateResponse("importar_datos.html", {
                "request": request,
                "current_user": user,
                "resultado_usuarios": {
                    "exitosos": 0,
                    "errores": [{"fila": 0, "error": mensaje_error}]
                }
            })
        
        # Guardar archivo temporalmente
        temp_path = f"uploads/temp_{archivo.filename}"
        with open(temp_path, "wb") as buffer:
            shutil.copyfileobj(archivo.file, buffer)
        
        # Importar usuarios
        resultado = excel_import.importar_usuarios_desde_excel(temp_path, db)
        
        # Eliminar archivo temporal
        os.remove(temp_path)
        
        return templates.TemplateResponse("importar_datos.html", {
            "request": request,
            "current_user": user,
            "resultado_usuarios": resultado
        })
        
    except Exception as e:
        return templates.TemplateResponse("importar_datos.html", {
            "request": request,
            "current_user": user,
            "resultado_usuarios": {
                "exitosos": 0,
                "errores": [{"fila": 0, "error": f"Error procesando archivo: {str(e)}"}]
            }
        })


@app.post("/importar-prospectos")
async def importar_prospectos(
    request: Request,
    archivo: UploadFile = File(...),
    db: Session = Depends(database.get_db),
    user: models.Usuario = Depends(require_admin)
):
    """Procesa la importación de prospectos desde Excel"""
    try:
        # Validar archivo
        es_valido, mensaje_error = excel_import.validar_archivo_excel(archivo)
        if not es_valido:
            return templates.TemplateResponse("importar_datos.html", {
                "request": request,
                "current_user": user,
                "resultado_prospectos": {
                    "exitosos": 0,
                    "errores": [{"fila": 0, "error": mensaje_error}],
                    "recurrentes": 0
                }
            })
        
        # Guardar archivo temporalmente
        temp_path = f"uploads/temp_{archivo.filename}"
        with open(temp_path, "wb") as buffer:
            shutil.copyfileobj(archivo.file, buffer)
        
        # Importar prospectos
        resultado = excel_import.importar_prospectos_desde_excel(temp_path, db)
        
        # Eliminar archivo temporal
        os.remove(temp_path)
        
        return templates.TemplateResponse("importar_datos.html", {
            "request": request,
            "current_user": user,
            "resultado_prospectos": resultado
        })
        
    except Exception as e:
        return templates.TemplateResponse("importar_datos.html", {
            "request": request,
            "current_user": user,
            "resultado_prospectos": {
                "exitosos": 0,
                "errores": [{"fila": 0, "error": f"Error procesando archivo: {str(e)}"}],
                "recurrentes": 0
            }
        })


@app.get("/descargar-plantilla/{tipo}")
async def descargar_plantilla(
    tipo: str,
    user: models.Usuario = Depends(require_admin)
):
    """Descarga plantilla Excel de ejemplo"""
    if tipo == "usuarios":
        file_path = "static/plantillas/plantilla_usuarios.xlsx"
        filename = "plantilla_usuarios.xlsx"
    elif tipo == "prospectos":
        file_path = "static/plantillas/plantilla_prospectos.xlsx"
        filename = "plantilla_prospectos.xlsx"
    else:
        raise HTTPException(status_code=404, detail="Plantilla no encontrada")
    
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="Archivo de plantilla no encontrado")
    
    return StreamingResponse(
        open(file_path, "rb"),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# Dashboard principal con filtros de fecha - VERSIÓN CORREGIDA
@app.get("/dashboard", response_class=HTMLResponse)
async def dashboard(
    request: Request,
    periodo: str = Query("mes"),  # dia, semana, mes, año, personalizado
    fecha_inicio: str = Query(None),
    fecha_fin: str = Query(None),
    db: Session = Depends(database.get_db)
):
    user = await get_current_user(request, db)
    
    if not user:
        return RedirectResponse(url="/", status_code=303)
    
    try:
        # Determinar el rango de fechas según el periodo seleccionado
        fecha_inicio_obj, fecha_fin_obj = calcular_rango_fechas(periodo, fecha_inicio, fecha_fin)
        
        print(f"📊 Calculando estadísticas para periodo: {periodo}")
        print(f"📅 Rango: {fecha_inicio_obj} a {fecha_fin_obj}")
        
        # Convertir a datetime para consultas
        fecha_inicio_dt = datetime.combine(fecha_inicio_obj, datetime.min.time())
        fecha_fin_dt = datetime.combine(fecha_fin_obj, datetime.max.time())
        
        # Inicializar todas las variables
        total_prospectos = prospectos_con_datos = prospectos_sin_datos = 0
        clientes_sin_asignar = clientes_asignados = destinos_count = ventas_count = 0
        prospectos_nuevos = prospectos_seguimiento = prospectos_cotizados = prospectos_ganados = prospectos_perdidos = ventas_canceladas = 0
        destinos_populares = []
        conversion_agentes = []
        
        # Estadísticas básicas
        if user.tipo_usuario in [TipoUsuario.ADMINISTRADOR.value, TipoUsuario.SUPERVISOR.value]:
            print("👨‍💼 Usuario es Admin/Supervisor - mostrando estadísticas generales")
            
            # ✅ CORREGIDO: Total de prospectos en el periodo (NO filtrado por estado)
            total_prospectos = db.query(models.Prospecto).filter(
                models.Prospecto.fecha_registro >= fecha_inicio_dt,
                models.Prospecto.fecha_registro <= fecha_fin_dt
            ).count()
            print(f"📈 Total prospectos: {total_prospectos}")
            
            # ✅ NUEVO: Prospectos con datos completos
            prospectos_con_datos = db.query(models.Prospecto).filter(
                models.Prospecto.tiene_datos_completos == True,
                models.Prospecto.fecha_registro >= fecha_inicio_dt,
                models.Prospecto.fecha_registro <= fecha_fin_dt
            ).count()
            print(f"📝 Prospectos con datos: {prospectos_con_datos}")
            
            # ✅ NUEVO: Prospectos sin datos (solo teléfono)
            prospectos_sin_datos = db.query(models.Prospecto).filter(
                models.Prospecto.tiene_datos_completos == False,
                models.Prospecto.fecha_registro >= fecha_inicio_dt,
                models.Prospecto.fecha_registro <= fecha_fin_dt
            ).count()
            print(f"📱 Prospectos sin datos: {prospectos_sin_datos}")
            
            # Clientes nuevos sin asignar en el periodo
            clientes_sin_asignar = db.query(models.Prospecto).filter(
                models.Prospecto.estado == EstadoProspecto.NUEVO.value,
                models.Prospecto.agente_asignado_id == None,
                models.Prospecto.fecha_registro >= fecha_inicio_dt,
                models.Prospecto.fecha_registro <= fecha_fin_dt
            ).count()
            print(f"🆕 Clientes sin asignar: {clientes_sin_asignar}")
            
            # Clientes asignados en el periodo (cualquier estado)
            clientes_asignados = db.query(models.Prospecto).filter(
                models.Prospecto.agente_asignado_id != None,
                models.Prospecto.fecha_registro >= fecha_inicio_dt,
                models.Prospecto.fecha_registro <= fecha_fin_dt
            ).count()
            print(f"📅 Clientes asignados: {clientes_asignados}")
            
            # Destinos registrados en el periodo
            destinos_query = db.query(models.Prospecto.destino).filter(
                models.Prospecto.fecha_registro >= fecha_inicio_dt,
                models.Prospecto.fecha_registro <= fecha_fin_dt,
                models.Prospecto.destino.isnot(None),
                models.Prospecto.destino != ''
            ).distinct().all()
            destinos_count = len(destinos_query)
            print(f"🌍 Destinos registrados: {destinos_count}")
            
            # Ventas registradas en el periodo
            ventas_count = db.query(models.Prospecto).filter(
                models.Prospecto.estado == EstadoProspecto.GANADO.value,
                models.Prospecto.fecha_registro >= fecha_inicio_dt,
                models.Prospecto.fecha_registro <= fecha_fin_dt
            ).count()
            print(f"💰 Ventas: {ventas_count}")
            
            # Destinos más solicitados en el periodo
            destinos_populares = db.query(
                models.Prospecto.destino,
                func.count(models.Prospecto.id).label('count')
            ).filter(
                models.Prospecto.fecha_registro >= fecha_inicio_dt,
                models.Prospecto.fecha_registro <= fecha_fin_dt,
                models.Prospecto.destino.isnot(None),
                models.Prospecto.destino != ''
            ).group_by(models.Prospecto.destino).order_by(func.count(models.Prospecto.id).desc()).limit(5).all()
            print(f"🏆 Destinos populares: {len(destinos_populares)}")
            
            # Estadísticas por estado en el periodo
            prospectos_nuevos = db.query(models.Prospecto).filter(
                models.Prospecto.estado == EstadoProspecto.NUEVO.value,
                models.Prospecto.fecha_registro >= fecha_inicio_dt,
                models.Prospecto.fecha_registro <= fecha_fin_dt
            ).count()
            
            prospectos_seguimiento = db.query(models.Prospecto).filter(
                models.Prospecto.estado == EstadoProspecto.EN_SEGUIMIENTO.value,
                models.Prospecto.fecha_registro >= fecha_inicio_dt,
                models.Prospecto.fecha_registro <= fecha_fin_dt
            ).count()
            
            # Cotizados: contar prospectos en estado COTIZADO actualmente
            prospectos_cotizados = db.query(models.Prospecto).filter(
                models.Prospecto.estado == EstadoProspecto.COTIZADO.value,
                models.Prospecto.fecha_registro >= fecha_inicio_dt,
                models.Prospecto.fecha_registro <= fecha_fin_dt
            ).count()
            
            # Ganados, Perdidos y Canceladas: contar estado ACTUAL
            prospectos_ganados = db.query(models.Prospecto).filter(
                models.Prospecto.estado == EstadoProspecto.GANADO.value,
                models.Prospecto.fecha_registro >= fecha_inicio_dt,
                models.Prospecto.fecha_registro <= fecha_fin_dt
            ).count()
            
            prospectos_perdidos = db.query(models.Prospecto).filter(
                models.Prospecto.estado == EstadoProspecto.CERRADO_PERDIDO.value,
                models.Prospecto.fecha_registro >= fecha_inicio_dt,
                models.Prospecto.fecha_registro <= fecha_fin_dt
            ).count()
            
            ventas_canceladas = db.query(models.Prospecto).filter(
                models.Prospecto.estado == EstadoProspecto.VENTA_CANCELADA.value,
                models.Prospecto.fecha_registro >= fecha_inicio_dt,
                models.Prospecto.fecha_registro <= fecha_fin_dt
            ).count()
            
            print(f"📊 Estados - Nuevos: {prospectos_nuevos}, Seguimiento: {prospectos_seguimiento}, Cotizados: {prospectos_cotizados}, Ganados: {prospectos_ganados}, Perdidos: {prospectos_perdidos}, Canceladas: {ventas_canceladas}")
            
            # Conversión por agente en el periodo
            conversion_agentes = []
            agentes_con_prospectos = db.query(
                models.Usuario.id,
                models.Usuario.username
            ).filter(
                models.Usuario.tipo_usuario == TipoUsuario.AGENTE.value,
                models.Usuario.activo == 1  # ✅ Solo agentes activos
            ).all()
            
            for agente in agentes_con_prospectos:
                total_agente = db.query(models.Prospecto).filter(
                    models.Prospecto.agente_asignado_id == agente.id,
                    models.Prospecto.fecha_registro >= fecha_inicio_dt,
                    models.Prospecto.fecha_registro <= fecha_fin_dt
                ).count()
                
                ganados_agente = db.query(models.HistorialEstado).filter(
                    models.HistorialEstado.usuario_id == agente.id,
                    models.HistorialEstado.estado_nuevo == EstadoProspecto.GANADO.value,
                    models.HistorialEstado.fecha_cambio >= fecha_inicio_dt,
                    models.HistorialEstado.fecha_cambio <= fecha_fin_dt
                ).count()
                
                cotizados_agente = db.query(models.EstadisticaCotizacion).filter(
                    models.EstadisticaCotizacion.agente_id == agente.id,
                    models.EstadisticaCotizacion.fecha_cotizacion >= fecha_inicio_obj,
                    models.EstadisticaCotizacion.fecha_cotizacion <= fecha_fin_obj
                ).count()

                conversion_agentes.append({
                    'id': agente.id,
                    'username': agente.username,
                    'total_prospectos': total_agente,
                    'cotizados': cotizados_agente,
                    'ganados': ganados_agente
                })
            
            print(f"👥 Conversión agentes: {len(conversion_agentes)} agentes")
            
        else:
            print("👤 Usuario es Agente - mostrando estadísticas personales")
            
            # Estadísticas para agente (solo sus datos) en el periodo
            total_prospectos = db.query(models.Prospecto).filter(
                models.Prospecto.agente_asignado_id == user.id,
                models.Prospecto.fecha_registro >= fecha_inicio_dt,
                models.Prospecto.fecha_registro <= fecha_fin_dt
            ).count()
            print(f"📈 Total prospectos agente: {total_prospectos}")
            
            # ✅ AGREGADO: Prospectos con datos completos para agente
            prospectos_con_datos = db.query(models.Prospecto).filter(
                models.Prospecto.agente_asignado_id == user.id,
                models.Prospecto.tiene_datos_completos == True,
                models.Prospecto.fecha_registro >= fecha_inicio_dt,
                models.Prospecto.fecha_registro <= fecha_fin_dt
            ).count()
            print(f"📝 Prospectos con datos agente: {prospectos_con_datos}")
            
            # ✅ AGREGADO: Prospectos sin datos para agente
            prospectos_sin_datos = db.query(models.Prospecto).filter(
                models.Prospecto.agente_asignado_id == user.id,
                models.Prospecto.tiene_datos_completos == False,
                models.Prospecto.fecha_registro >= fecha_inicio_dt,
                models.Prospecto.fecha_registro <= fecha_fin_dt
            ).count()
            print(f"📱 Prospectos sin datos agente: {prospectos_sin_datos}")
            
            # Clientes asignados al agente en el periodo
            clientes_asignados = db.query(models.Prospecto).filter(
                models.Prospecto.agente_asignado_id == user.id,
                models.Prospecto.fecha_registro >= fecha_inicio_dt,
                models.Prospecto.fecha_registro <= fecha_fin_dt
            ).count()
            print(f"📅 Clientes asignados agente: {clientes_asignados}")
            
            # Destinos registrados por el agente en el periodo
            destinos_query = db.query(models.Prospecto.destino).filter(
                models.Prospecto.agente_asignado_id == user.id,
                models.Prospecto.fecha_registro >= fecha_inicio_dt,
                models.Prospecto.fecha_registro <= fecha_fin_dt,
                models.Prospecto.destino.isnot(None),
                models.Prospecto.destino != ''
            ).distinct().all()
            destinos_count = len(destinos_query)
            print(f"🌍 Destinos registrados agente: {destinos_count}")
            
            # Ventas del agente en el periodo
            # Ventas del agente en el periodo (Basado en historial de cambios)
            ventas_count = db.query(models.HistorialEstado).filter(
                models.HistorialEstado.usuario_id == user.id,
                models.HistorialEstado.estado_nuevo == EstadoProspecto.GANADO.value,
                models.HistorialEstado.fecha_cambio >= fecha_inicio_dt,
                models.HistorialEstado.fecha_cambio <= fecha_fin_dt
            ).count()
            print(f"💰 Ventas agente: {ventas_count}")
            
            # Destinos más solicitados por el agente en el periodo
            destinos_populares = db.query(
                models.Prospecto.destino,
                func.count(models.Prospecto.id).label('count')
            ).filter(
                models.Prospecto.agente_asignado_id == user.id,
                models.Prospecto.fecha_registro >= fecha_inicio_dt,
                models.Prospecto.fecha_registro <= fecha_fin_dt,
                models.Prospecto.destino.isnot(None),
                models.Prospecto.destino != ''
            ).group_by(models.Prospecto.destino).order_by(func.count(models.Prospecto.id).desc()).limit(5).all()
            print(f"🏆 Destinos populares agente: {len(destinos_populares)}")
            
            # Para agente, no mostrar estos datos generales
            clientes_sin_asignar = 0
            
            # Estadísticas por estado para agente en el periodo
            prospectos_nuevos = db.query(models.Prospecto).filter(
                models.Prospecto.agente_asignado_id == user.id,
                models.Prospecto.estado == EstadoProspecto.NUEVO.value,
                models.Prospecto.fecha_registro >= fecha_inicio_dt,
                models.Prospecto.fecha_registro <= fecha_fin_dt
            ).count()
            prospectos_seguimiento = db.query(models.HistorialEstado).filter(
                models.HistorialEstado.usuario_id == user.id,
                models.HistorialEstado.estado_nuevo == EstadoProspecto.EN_SEGUIMIENTO.value,
                models.HistorialEstado.fecha_cambio >= fecha_inicio_dt,
                models.HistorialEstado.fecha_cambio <= fecha_fin_dt
            ).count()
            
            # Nota: Cotizados ya usa EstadisticaCotizacion (correcto)
            prospectos_cotizados = db.query(models.EstadisticaCotizacion).filter(
                models.EstadisticaCotizacion.agente_id == user.id,
                models.EstadisticaCotizacion.fecha_cotizacion >= fecha_inicio_obj,
                models.EstadisticaCotizacion.fecha_cotizacion <= fecha_fin_obj
            ).count()
            
            prospectos_ganados = db.query(models.HistorialEstado).filter(
                models.HistorialEstado.usuario_id == user.id,
                models.HistorialEstado.estado_nuevo == EstadoProspecto.GANADO.value,
                models.HistorialEstado.fecha_cambio >= fecha_inicio_dt,
                models.HistorialEstado.fecha_cambio <= fecha_fin_dt
            ).count()
            
            prospectos_perdidos = db.query(models.HistorialEstado).filter(
                models.HistorialEstado.usuario_id == user.id,
                models.HistorialEstado.estado_nuevo == EstadoProspecto.CERRADO_PERDIDO.value,
                models.HistorialEstado.fecha_cambio >= fecha_inicio_dt,
                models.HistorialEstado.fecha_cambio <= fecha_fin_dt
            ).count()
            
            print(f"📊 Estados agente - Nuevos: {prospectos_nuevos}, Seguimiento: {prospectos_seguimiento}, Cotizados: {prospectos_cotizados}, Ganados: {prospectos_ganados}, Perdidos: {prospectos_perdidos}")
            
            conversion_agentes = []
        
    except Exception as e:
        print(f"❌ Error grave calculando estadísticas: {e}")
        import traceback
        traceback.print_exc()
        # Inicializar todas las variables con valores por defecto
        total_prospectos = prospectos_con_datos = prospectos_sin_datos = 0
        clientes_sin_asignar = clientes_asignados = destinos_count = ventas_count = 0
        prospectos_nuevos = prospectos_seguimiento = prospectos_cotizados = prospectos_ganados = prospectos_perdidos = 0
        destinos_populares = []
        conversion_agentes = []
        fecha_inicio_obj = date.today()
        fecha_fin_obj = date.today()
    
    return templates.TemplateResponse("dashboard.html", {
        "request": request,
        "current_user": user,
        "today": date.today().strftime("%d/%m/%Y"),
        
        # Filtros activos
        "periodo_activo": periodo,
        "fecha_inicio_activa": fecha_inicio,
        "fecha_fin_activa": fecha_fin,
        "fecha_inicio_formateada": fecha_inicio_obj.strftime("%d/%m/%Y") if fecha_inicio_obj else "",
        "fecha_fin_formateada": fecha_fin_obj.strftime("%d/%m/%Y") if fecha_fin_obj else "",
        
        # Estadísticas principales
        "total_prospectos": total_prospectos,
        "prospectos_con_datos": prospectos_con_datos,
        "prospectos_sin_datos": prospectos_sin_datos,
        "clientes_sin_asignar": clientes_sin_asignar,
        "clientes_asignados": clientes_asignados,
        "destinos_count": destinos_count,
        "ventas_count": ventas_count,
        
        # Estadísticas por estado
        "prospectos_nuevos": prospectos_nuevos,
        "prospectos_seguimiento": prospectos_seguimiento,
        "prospectos_cotizados": prospectos_cotizados,
        "prospectos_ganados": prospectos_ganados,
        "prospectos_perdidos": prospectos_perdidos,
        "ventas_canceladas": ventas_canceladas,
        
        # Datos para gráficos
        "destinos_populares": destinos_populares,
        "conversion_agentes": conversion_agentes
    })


# ========== GESTIÓN DE PROSPECTOS (ACTUALIZADO) ==========

@app.get("/prospectos", response_class=HTMLResponse)
async def listar_prospectos(
    request: Request,
    destino: str = Query(None),
    telefono: str = Query(None),
    medio_ingreso_id: str = Query(None),
    agente_asignado_id: str = Query(None),
    estado: str = Query(None),
    busqueda_global: str = Query(None),
    fecha_inicio: str = Query(None),  # ✅ NUEVO: Filtro de fecha inicio
    fecha_fin: str = Query(None),  # ✅ NUEVO: Filtro de fecha fin
    page: int = Query(1, ge=1),  # ✅ Paginación: Página actual
    limit: int = Query(10, ge=1, le=100),  # ✅ Paginación: Registros por página
    db: Session = Depends(database.get_db)
):
    user = await get_current_user(request, db)
    
    if not user:
        return RedirectResponse(url="/", status_code=303)
    
    # ✅ DETERMINAR RANGO DE FECHAS (DEFAULT: MES ACTUAL)
    if not fecha_inicio or not fecha_fin:
        hoy = date.today()
        fecha_inicio_date = date(hoy.year, hoy.month, 1)
        ultimo_dia = calendar.monthrange(hoy.year, hoy.month)[1]
        fecha_fin_date = date(hoy.year, hoy.month, ultimo_dia)
        # Convertir a string para la vista
        fecha_inicio = fecha_inicio_date.strftime("%Y-%m-%d")
        fecha_fin = fecha_fin_date.strftime("%Y-%m-%d")
    else:
        fecha_inicio_date = datetime.strptime(fecha_inicio, "%Y-%m-%d").date()
        fecha_fin_date = datetime.strptime(fecha_fin, "%Y-%m-%d").date()
    
    # ✅ CONSTRUIR QUERY BASE SEGÚN ROL (SIN FILTRO DE ESTADO INICIAL)
    if user.tipo_usuario == TipoUsuario.AGENTE.value:
        query = db.query(models.Prospecto).filter(
            models.Prospecto.agente_asignado_id == user.id
        )
    else:
        query = db.query(models.Prospecto)
    
    # ✅ LÓGICA DE FILTROS POR DEFECTO Y EXPLÍCITOS
    filtros_aplicados = False

    if estado:
        filtros_aplicados = True
        if estado == "todos":
            pass
        else:
            query = query.filter(models.Prospecto.estado == estado)
    
    if agente_asignado_id:
        filtros_aplicados = True
        if agente_asignado_id == "todos":
            pass
        elif agente_asignado_id == "sin_asignar":
            query = query.filter(models.Prospecto.agente_asignado_id == None)
        else:
            query = query.filter(models.Prospecto.agente_asignado_id == int(agente_asignado_id))

    # ✅ APLICAR VALORES POR DEFECTO SI NO HAY ELEMENTOS DE FILTRO ESPECÍFICOS EN LA URL
    # Se considera "filtro activo" si el usuario envió algún parámetro en la URL
    # Pero para simplificar, si 'estado' y 'agente_asignado_id' son None, aplicamos defaults
    
    if estado is None and agente_asignado_id is None:
        if user.tipo_usuario == TipoUsuario.AGENTE.value:
            # Agente: Nuevo, Seguimiento, Cotizado
            query = query.filter(models.Prospecto.estado.in_([
                EstadoProspecto.NUEVO.value,
                EstadoProspecto.EN_SEGUIMIENTO.value,
                EstadoProspecto.COTIZADO.value
            ]))
        else:
            # Admin/Supervisor: Nuevo Y Sin Asignar
            query = query.filter(models.Prospecto.estado == EstadoProspecto.NUEVO.value)
            query = query.filter(models.Prospecto.agente_asignado_id == None)
            
            # Ajustar valores para que se reflejen en la vista
            estado = EstadoProspecto.NUEVO.value
            agente_asignado_id = "sin_asignar"
    
    # ✅ FILTRO POR RANGO DE FECHAS (fecha_registro)
    query = query.filter(
        models.Prospecto.fecha_registro >= fecha_inicio_date,
        models.Prospecto.fecha_registro <= fecha_fin_date
    )

    # ✅ FILTRO DE BÚSQUEDA GLOBAL
    if busqueda_global:
        search_term = f"%{busqueda_global}%"
        query = query.filter(
            or_(
                models.Prospecto.nombre.ilike(search_term),
                models.Prospecto.apellido.ilike(search_term),
                models.Prospecto.telefono.ilike(search_term),
                models.Prospecto.telefono_secundario.ilike(search_term),
                models.Prospecto.correo_electronico.ilike(search_term),
                models.Prospecto.destino.ilike(search_term),
                models.Prospecto.ciudad_origen.ilike(search_term),
                models.Prospecto.observaciones.ilike(search_term)
            )
        )
        print(f"🔍 Aplicando búsqueda global: {busqueda_global}")
    
    # ✅ FILTRO POR TELÉFONO
    if telefono:
        telefono_term = f"%{telefono}%"
        query = query.filter(
            or_(
                models.Prospecto.telefono.ilike(telefono_term),
                models.Prospecto.telefono_secundario.ilike(telefono_term)
            )
        )
    
    # Otros filtros
    if destino:
        query = query.filter(models.Prospecto.destino.ilike(f"%{destino}%"))
    
    if medio_ingreso_id and medio_ingreso_id != "todos":
        query = query.filter(models.Prospecto.medio_ingreso_id == int(medio_ingreso_id))
    
    # ✅ ORDENAMIENTO: Del más nuevo al más antiguo
    # Se ordena después de todos los filtros y antes de la paginación
    query = query.order_by(models.Prospecto.fecha_registro.desc())
    
    # ✅ PAGINACIÓN
    total_registros = query.count()
    total_pages = (total_registros + limit - 1) // limit
    
    # Asegurar que la página solicitada sea válida
    if page > total_pages and total_pages > 0:
        page = total_pages
    
    offset = (page - 1) * limit
    prospectos = query.offset(offset).limit(limit).all()
    
    # Obtener datos para filtros
    agentes = db.query(models.Usuario).filter(
        models.Usuario.tipo_usuario == TipoUsuario.AGENTE.value,
        models.Usuario.activo == 1  # ✅ Solo agentes activos
    ).all()
    
    medios_ingreso = db.query(models.MedioIngreso).all()
    
    return templates.TemplateResponse("prospectos.html", {
        "request": request,
        "prospectos": prospectos,
        "current_user": user,
        "agentes": agentes,
        "medios_ingreso": medios_ingreso,
        "estados_prospecto": [estado.value for estado in EstadoProspecto],
        "filtros_activos": {
            "destino": destino,
            "telefono": telefono,
            "medio_ingreso_id": medio_ingreso_id,
            "agente_asignado_id": agente_asignado_id,
            "estado": estado,
            "busqueda_global": busqueda_global
        },
        # Info Paginación
        "page": page,
        "limit": limit,
        "total_pages": total_pages,
        "total_registros": total_registros,
        # ✅ NUEVO: Filtros de fecha
        "fecha_inicio": fecha_inicio,
        "fecha_fin": fecha_fin
    })

@app.post("/prospectos")
async def crear_prospecto(
    request: Request,
    telefono: str = Form(...),
    indicativo_telefono: str = Form("57"),
    medio_ingreso_id: int = Form(...),
    nombre: str = Form(None),
    apellido: str = Form(None),
    correo_electronico: str = Form(None),
    ciudad_origen: str = Form(None),
    destino: str = Form(None),
    fecha_ida: str = Form(None),
    fecha_vuelta: str = Form(None),
    pasajeros_adultos: int = Form(1),
    pasajeros_ninos: int = Form(0),
    pasajeros_infantes: int = Form(0),
    observaciones: str = Form(None),
    empresa_segundo_titular: str = Form(None),  # ✅ NUEVO: Empresa o segundo titular
    telefono_secundario: str = Form(None),
    indicativo_telefono_secundario: str = Form("57"),
    forzar_nuevo: bool = Form(False),
    agente_asignado_id: int = Form(None),
    db: Session = Depends(database.get_db)
):
    user = await get_current_user(request, db)
    if not user:
        return RedirectResponse(url="/", status_code=303)
    
    try:
        # ✅ VALIDACIÓN DE INDICATIVOS
        if not indicativo_telefono.isdigit() or len(indicativo_telefono) > 4:
            return RedirectResponse(url="/prospectos?error=Indicativo principal inválido. Solo números, máximo 4 dígitos", status_code=303)
        
        if indicativo_telefono_secundario and (not indicativo_telefono_secundario.isdigit() or len(indicativo_telefono_secundario) > 4):
            return RedirectResponse(url="/prospectos?error=Indicativo secundario inválido. Solo números, máximo 4 dígitos", status_code=303)

        # ✅ DETECCIÓN MEJORADA: OBTENER TODOS LOS REGISTROS DEL CLIENTE
        # Buscar por teléfono principal
        clientes_existentes_principal = db.query(models.Prospecto).filter(
            or_(
                models.Prospecto.telefono == telefono,
                models.Prospecto.telefono_secundario == telefono
            )
        ).all()

        # Buscar por teléfono secundario (si existe)
        clientes_existentes_secundario = []
        if telefono_secundario:
            clientes_existentes_secundario = db.query(models.Prospecto).filter(
                or_(
                    models.Prospecto.telefono == telefono_secundario,
                    models.Prospecto.telefono_secundario == telefono_secundario
                )
            ).all()

        # Combinar resultados únicos
        clientes_existentes_set = set()
        todos_clientes_existentes = []

        # Agregar de principal
        for cliente in clientes_existentes_principal:
            if cliente.id not in clientes_existentes_set:
                clientes_existentes_set.add(cliente.id)
                todos_clientes_existentes.append(cliente)

        # Agregar de secundario
        for cliente in clientes_existentes_secundario:
            if cliente.id not in clientes_existentes_set:
                clientes_existentes_set.add(cliente.id)
                todos_clientes_existentes.append(cliente)

        # ✅ ORDENAR POR fecha_registro (nombre correcto en tu modelo)
        todos_clientes_existentes.sort(key=lambda x: x.fecha_registro, reverse=True)

        # Usar el más reciente como "cliente principal" para compatibilidad
        cliente_existente_principal = todos_clientes_existentes[0] if todos_clientes_existentes else None
        
        if todos_clientes_existentes and not forzar_nuevo:
            # Obtener datos de TODOS los registros del cliente
            todos_ids = [c.id for c in todos_clientes_existentes]
            
            interacciones_previas = db.query(models.Interaccion).filter(
                models.Interaccion.prospecto_id.in_(todos_ids)
            ).count()
            
            documentos_previos = db.query(models.Documento).filter(
                models.Documento.prospecto_id.in_(todos_ids)
            ).count()
            
            # Obtener últimas interacciones de TODOS los registros
            ultimas_interacciones = db.query(models.Interaccion).filter(
                models.Interaccion.prospecto_id.in_(todos_ids)
            ).order_by(models.Interaccion.fecha_creacion.desc()).limit(5).all()
            
            # Preparar datos para el template
            nuevos_datos = {
                "telefono": telefono,
                "indicativo_telefono": indicativo_telefono,
                "nombre": nombre,
                "apellido": apellido,
                "correo_electronico": correo_electronico,
                "ciudad_origen": ciudad_origen,
                "destino": destino,
                "fecha_ida": fecha_ida,
                "fecha_vuelta": fecha_vuelta,
                "pasajeros_adultos": pasajeros_adultos,
                "pasajeros_ninos": pasajeros_ninos,
                "pasajeros_infantes": pasajeros_infantes,
                "medio_ingreso_id": medio_ingreso_id,
                "telefono_secundario": telefono_secundario,
                "indicativo_telefono_secundario": indicativo_telefono_secundario,
                "observaciones": observaciones
            }
            
            # ✅ AGREGAR: Obtener lista de agentes para el select
            agentes = db.query(models.Usuario).filter(
                models.Usuario.tipo_usuario == TipoUsuario.AGENTE.value
            ).all()
            
            # Renderizar template de confirmación
            return templates.TemplateResponse("confirmar_cliente_existente.html", {
                "request": request,
                "cliente_existente_principal": cliente_existente_principal,
                "registros_previos": todos_clientes_existentes,
                "interacciones_previas": interacciones_previas,
                "documentos_previos": documentos_previos,
                "ultimas_interacciones": ultimas_interacciones,
                "nuevos_datos": nuevos_datos,
                "agentes": agentes  # ✅ Variable que faltaba
            })
        
        # ✅ DETERMINAR AGENTE ASIGNADO
        # 1. Si se especificó en el formulario, usar ese
        # 2. Si el usuario es agente, asignarse a sí mismo
        # 3. Si es admin/supervisor y no especificó, dejar sin asignar
        agente_final_id = None
        if agente_asignado_id and agente_asignado_id != 0:
            # Verificar que el agente exista
            agente = db.query(models.Usuario).filter(
                models.Usuario.id == agente_asignado_id,
                models.Usuario.tipo_usuario == TipoUsuario.AGENTE.value
            ).first()
            if agente:
                agente_final_id = agente_asignado_id
        elif user.tipo_usuario == TipoUsuario.AGENTE.value:
            agente_final_id = user.id

        # ✅ DETERMINAR DATOS PARA EL NUEVO REGISTRO
        # Prioridad: 1) Datos del formulario, 2) Datos del cliente más reciente, 3) Ninguno
        nombre_final = nombre
        apellido_final = apellido
        email_final = correo_electronico

        # Si no se proporcionaron en el formulario, intentar usar del cliente más reciente
        if not nombre_final and todos_clientes_existentes:
            for cliente in todos_clientes_existentes:
                if cliente.nombre:
                    nombre_final = cliente.nombre
                    break

        if not apellido_final and todos_clientes_existentes:
            for cliente in todos_clientes_existentes:
                if cliente.apellido:
                    apellido_final = cliente.apellido
                    break

        if not email_final and todos_clientes_existentes:
            for cliente in todos_clientes_existentes:
                if cliente.correo_electronico:
                    email_final = cliente.correo_electronico
                    break

        # ✅ NUEVO: Copiar campos adicionales de clientes ganados
        fecha_nacimiento_final = None
        numero_identificacion_final = None
        direccion_final = None
        
        if todos_clientes_existentes:
            for cliente in todos_clientes_existentes:
                if cliente.fecha_nacimiento and not fecha_nacimiento_final:
                    fecha_nacimiento_final = cliente.fecha_nacimiento
                if cliente.numero_identificacion and not numero_identificacion_final:
                    numero_identificacion_final = cliente.numero_identificacion
                if cliente.direccion and not direccion_final:
                    direccion_final = cliente.direccion

        # ✅ CREAR NUEVO PROSPECTO CON INDICATIVOS Y DATOS MEJORADOS
        fecha_ida_date = normalizar_fecha_input(fecha_ida)
        fecha_vuelta_date = normalizar_fecha_input(fecha_vuelta)
        
        # ✅ NORMALIZAR DATOS ANTES DE GUARDAR
        telefono_normalizado = normalizar_numero(telefono)
        telefono_secundario_normalizado = normalizar_numero(telefono_secundario) if telefono_secundario else None
        nombre_normalizado = normalizar_texto_mayusculas(nombre_final)
        apellido_normalizado = normalizar_texto_mayusculas(apellido_final)
        ciudad_origen_normalizada = normalizar_texto_mayusculas(ciudad_origen)
        destino_normalizado = normalizar_texto_mayusculas(destino)
        email_normalizado = normalizar_email(email_final)
        empresa_segundo_titular_normalizado = normalizar_texto_mayusculas(empresa_segundo_titular)  # ✅ NUEVO
        # observaciones NO se normalizan (mantener formato original)
        
        # Determinar si es cliente recurrente
        cliente_recurrente = len(todos_clientes_existentes) > 0

        prospecto = models.Prospecto(
            nombre=nombre_normalizado,  # ✅ NORMALIZADO A MAYÚSCULAS
            apellido=apellido_normalizado,  # ✅ NORMALIZADO A MAYÚSCULAS
            correo_electronico=email_normalizado,  # ✅ NORMALIZADO A MINÚSCULAS
            telefono=telefono_normalizado,  # ✅ SOLO NÚMEROS
            indicativo_telefono=indicativo_telefono,
            telefono_secundario=telefono_secundario_normalizado,  # ✅ SOLO NÚMEROS
            indicativo_telefono_secundario=indicativo_telefono_secundario,
            ciudad_origen=ciudad_origen_normalizada,  # ✅ NORMALIZADO A MAYÚSCULAS
            destino=destino_normalizado,  # ✅ NORMALIZADO A MAYÚSCULAS
            fecha_ida=fecha_ida_date,
            fecha_vuelta=fecha_vuelta_date,
            pasajeros_adultos=pasajeros_adultos,
            pasajeros_ninos=pasajeros_ninos,
            pasajeros_infantes=pasajeros_infantes,
            medio_ingreso_id=medio_ingreso_id,
            observaciones=observaciones,  # ✅ SIN NORMALIZAR (mantener original)
            agente_asignado_id=agente_final_id,  # ✅ USAR AGENTE DETERMINADO
            cliente_recurrente=cliente_recurrente,
            prospecto_original_id=todos_clientes_existentes[0].id if todos_clientes_existentes else None,
            # ✅ NUEVO: Copiar datos adicionales del cliente existente
            fecha_nacimiento=fecha_nacimiento_final,
            numero_identificacion=numero_identificacion_final,
            direccion=direccion_final,
            empresa_segundo_titular=empresa_segundo_titular_normalizado  # ✅ NUEVO
        )
        
        # ✅ VERIFICAR Y ASIGNAR DATOS COMPLETOS
        prospecto.verificar_datos_completos()
        
        db.add(prospecto)
        db.flush()  # Para obtener el ID antes del commit
        
        # ✅ GENERAR IDs ÚNICOS
        # 1. Reutilizar id_cliente si es cliente recurrente, sino generar nuevo
        if todos_clientes_existentes and todos_clientes_existentes[0].id_cliente:
            prospecto.id_cliente = todos_clientes_existentes[0].id_cliente
            print(f"♻️ Reutilizando id_cliente: {prospecto.id_cliente}")
        else:
            prospecto.generar_id_cliente()
            print(f"✅ Nuevo id_cliente generado: {prospecto.id_cliente}")
        
        # 2. Siempre generar nuevo id_solicitud (único por caso/viaje)
        prospecto.generar_id_solicitud()
        print(f"✅ Nuevo id_solicitud generado: {prospecto.id_solicitud}")
        
        db.commit()

        
        # Registrar interacción automática si es recurrente
        if cliente_recurrente:
            interaccion = models.Interaccion(
                prospecto_id=prospecto.id,
                usuario_id=user.id,
                tipo_interaccion="sistema",
                descripcion=f"Cliente recurrente registrado. Teléfono: {telefono}. Registros previos: {len(todos_clientes_existentes)}",
                estado_anterior=cliente_existente_principal.estado if cliente_existente_principal else None,
                estado_nuevo=EstadoProspecto.NUEVO.value
            )
            db.add(interaccion)
            db.commit()
        
        mensaje = "Prospecto creado correctamente" + (" (Cliente recurrente)" if cliente_recurrente else "")
        return RedirectResponse(url=f"/prospectos?success={mensaje}", status_code=303)
    
    except Exception as e:
        db.rollback()
        print(f"❌ Error creating prospect: {e}")
        return RedirectResponse(url="/prospectos?error=Error al crear prospecto", status_code=303)



@app.get("/prospectos/{prospecto_id}/editar")
async def mostrar_editar_prospecto(
    request: Request,
    prospecto_id: int,
    db: Session = Depends(database.get_db)
):
    user = await get_current_user(request, db)
    if not user:
        return RedirectResponse(url="/", status_code=303)
    
    # Buscar prospecto
    prospecto = db.query(models.Prospecto).filter(models.Prospecto.id == prospecto_id).first()
    if not prospecto:
        return RedirectResponse(url="/prospectos?error=Prospecto no encontrado", status_code=303)
    
    # Verificar permisos: Agentes solo pueden editar sus propios prospectos
    if (user.tipo_usuario == TipoUsuario.AGENTE.value and 
        prospecto.agente_asignado_id != user.id):
        return RedirectResponse(url="/prospectos?error=No tiene permisos para editar este prospecto", status_code=303)
    
    # Obtener medios de ingreso para el dropdown
    medios_ingreso = db.query(models.MedioIngreso).all()
    
    return templates.TemplateResponse("editar_prospecto.html", {
        "request": request,
        "prospecto": prospecto,
        "medios_ingreso": medios_ingreso,
        "user": user
    })


@app.post("/prospectos/{prospecto_id}/editar")
async def editar_prospecto(
    request: Request,
    prospecto_id: int,
    telefono: str = Form(...),
    indicativo_telefono: str = Form("57"),
    medio_ingreso_id: int = Form(...),
    nombre: str = Form(None),
    apellido: str = Form(None),
    correo_electronico: str = Form(None),
    ciudad_origen: str = Form(None),
    destino: str = Form(None),
    fecha_ida: str = Form(None),
    fecha_vuelta: str = Form(None),
    pasajeros_adultos: int = Form(1),
    pasajeros_ninos: int = Form(0),
    pasajeros_infantes: int = Form(0),
    observaciones: str = Form(None),
    empresa_segundo_titular: str = Form(None),  # ✅ NUEVO: Empresa o segundo titular
    telefono_secundario: str = Form(None),
    indicativo_telefono_secundario: str = Form("57"),
    estado: str = Form(None),
    fecha_nacimiento: str = Form(None),
    numero_identificacion: str = Form(None),
    direccion: str = Form(None),  # ✅ NUEVO: Dirección para clientes ganados
    origen_solicitud: str = Form(None),  # Nuevo parámetro
    db: Session = Depends(database.get_db)
):
    user = await get_current_user(request, db)
    if not user:
        return RedirectResponse(url="/", status_code=303)
    
    try:
        # ✅ AGREGAR VALIDACIÓN DE INDICATIVOS (COMO EN CREAR_PROSPECTO)
        if not indicativo_telefono.isdigit() or len(indicativo_telefono) > 4:
            redirect_url = f"/prospectos/{prospecto_id}/seguimiento?error=Indicativo principal inválido" if origen_solicitud == "seguimiento" else "/prospectos?error=Indicativo principal inválido. Solo números, máximo 4 dígitos"
            return RedirectResponse(url=redirect_url, status_code=303)
        
        if indicativo_telefono_secundario and (not indicativo_telefono_secundario.isdigit() or len(indicativo_telefono_secundario) > 4):
            redirect_url = f"/prospectos/{prospecto_id}/seguimiento?error=Indicativo secundario inválido" if origen_solicitud == "seguimiento" else "/prospectos?error=Indicativo secundario inválido. Solo números, máximo 4 dígitos"
            return RedirectResponse(url=redirect_url, status_code=303)

        # Buscar prospecto
        prospecto = db.query(models.Prospecto).filter(models.Prospecto.id == prospecto_id).first()
        if not prospecto:
            redirect_url = f"/prospectos/{prospecto_id}/seguimiento?error=Prospecto no encontrado" if origen_solicitud == "seguimiento" else "/prospectos?error=Prospecto no encontrado"
            return RedirectResponse(url=redirect_url, status_code=303)
        
        # Verificar permisos: Agentes solo pueden editar sus propios prospectos
        if (user.tipo_usuario == TipoUsuario.AGENTE.value and 
            prospecto.agente_asignado_id != user.id):
            redirect_url = f"/prospectos/{prospecto_id}/seguimiento?error=No tiene permisos" if origen_solicitud == "seguimiento" else "/prospectos?error=No tiene permisos para editar este prospecto"
            return RedirectResponse(url=redirect_url, status_code=303)
        
        # Convertir fechas de string a date
        fecha_ida_date = normalizar_fecha_input(fecha_ida)
        fecha_vuelta_date = normalizar_fecha_input(fecha_vuelta)
        fecha_nacimiento_date = normalizar_fecha_input(fecha_nacimiento) if fecha_nacimiento else None
        
        # ✅ NORMALIZAR DATOS ANTES DE ACTUALIZAR
        telefono_normalizado = normalizar_numero(telefono)
        telefono_secundario_normalizado = normalizar_numero(telefono_secundario) if telefono_secundario else None
        nombre_normalizado = normalizar_texto_mayusculas(nombre)
        apellido_normalizado = normalizar_texto_mayusculas(apellido)
        ciudad_origen_normalizada = normalizar_texto_mayusculas(ciudad_origen)
        destino_normalizado = normalizar_texto_mayusculas(destino)
        email_normalizado = normalizar_email(correo_electronico)
        numero_identificacion_normalizado = normalizar_numero(numero_identificacion) if numero_identificacion else None
        direccion_normalizada = normalizar_texto_mayusculas(direccion)  # ✅ NUEVO
        empresa_segundo_titular_normalizado = normalizar_texto_mayusculas(empresa_segundo_titular)  # ✅ NUEVO
        # observaciones NO se normalizan (mantener formato original)
        
        # Validar cambio de estado a VENTA_CANCELADA
        if estado == EstadoProspecto.VENTA_CANCELADA.value:
            if prospecto.estado != EstadoProspecto.GANADO.value and prospecto.estado_anterior != EstadoProspecto.GANADO.value:
                redirect_url = f"/prospectos/{prospecto_id}/seguimiento?error=Solo se puede cancelar una venta que haya estado en estado GANADO" if origen_solicitud == "seguimiento" else "/prospectos?error=Solo se puede cancelar una venta que haya estado en estado GANADO"
                return RedirectResponse(url=redirect_url, status_code=303)
        
        # Guardar estado anterior si cambia
        estado_cambio_a_ganado = False
        if estado and estado != prospecto.estado:
            prospecto.estado_anterior = prospecto.estado
            prospecto.estado = estado
            # Detectar si cambió a ganado
            if estado == EstadoProspecto.GANADO.value:
                estado_cambio_a_ganado = True
        
        # Detectar si cambió fecha_ida
        fecha_ida_original = prospecto.fecha_ida
        fecha_ida_cambio = (fecha_ida_date != fecha_ida_original)
        
        # Actualizar datos del prospecto
        prospecto.nombre = nombre_normalizado  # ✅ NORMALIZADO A MAYÚSCULAS
        prospecto.apellido = apellido_normalizado  # ✅ NORMALIZADO A MAYÚSCULAS
        prospecto.correo_electronico = email_normalizado  # ✅ NORMALIZADO A MINÚSCULAS
        prospecto.telefono = telefono_normalizado  # ✅ SOLO NÚMEROS
        prospecto.indicativo_telefono = indicativo_telefono
        prospecto.telefono_secundario = telefono_secundario_normalizado  # ✅ SOLO NÚMEROS
        prospecto.indicativo_telefono_secundario = indicativo_telefono_secundario
        prospecto.ciudad_origen = ciudad_origen_normalizada  # ✅ NORMALIZADO A MAYÚSCULAS
        prospecto.destino = destino_normalizado  # ✅ NORMALIZADO A MAYÚSCULAS
        prospecto.fecha_ida = fecha_ida_date
        prospecto.fecha_vuelta = fecha_vuelta_date
        prospecto.pasajeros_adultos = pasajeros_adultos
        prospecto.pasajeros_ninos = pasajeros_ninos
        prospecto.pasajeros_infantes = pasajeros_infantes
        prospecto.medio_ingreso_id = medio_ingreso_id
        prospecto.observaciones = observaciones  # ✅ SIN NORMALIZAR (mantener original)
        prospecto.empresa_segundo_titular = empresa_segundo_titular_normalizado  # ✅ NUEVO
        
        # Campos adicionales para clientes GANADOS
        if prospecto.estado == EstadoProspecto.GANADO.value:
            prospecto.fecha_nacimiento = fecha_nacimiento_date
            prospecto.numero_identificacion = numero_identificacion_normalizado  # ✅ SOLO NÚMEROS
            prospecto.direccion = direccion_normalizada  # ✅ NUEVO: MAYÚSCULAS
        
        db.commit()
        
        # ✅ NUEVO: Crear notificaciones automáticas de viaje
        # Si cambió a ganado O si es ganado y cambió la fecha_ida
        if estado_cambio_a_ganado or (prospecto.estado == EstadoProspecto.GANADO.value and fecha_ida_cambio and fecha_ida_date):
            crear_notificaciones_viaje(prospecto, db)
        
        # Redirigir según origen
        if origen_solicitud == "seguimiento":
            return RedirectResponse(url=f"/prospectos/{prospecto_id}/seguimiento?success=Datos actualizados correctamente", status_code=303)
        else:
            return RedirectResponse(url="/prospectos?success=Prospecto actualizado correctamente", status_code=303)
    
    except Exception as e:
        db.rollback()
        print(f"❌ Error updating prospect: {e}")
        redirect_url = f"/prospectos/{prospecto_id}/seguimiento?error=Error al actualizar" if origen_solicitud == "seguimiento" else "/prospectos?error=Error al actualizar prospecto"
        return RedirectResponse(url=redirect_url, status_code=303)

@app.post("/prospectos/{prospecto_id}/eliminar")
async def eliminar_prospecto(
    request: Request,
    prospecto_id: int,
    db: Session = Depends(database.get_db)
):
    user = await get_current_user(request, db)
    if not user:
        return RedirectResponse(url="/", status_code=303)
    
    try:
        # Buscar prospecto
        prospecto = db.query(models.Prospecto).filter(models.Prospecto.id == prospecto_id).first()
        if not prospecto:
            return RedirectResponse(url="/prospectos?error=Prospecto no encontrado", status_code=303)
        
        # Verificar permisos: Agentes solo pueden eliminar sus propios prospectos
        # Admin/Supervisor pueden eliminar cualquier prospecto
        if (user.tipo_usuario == TipoUsuario.AGENTE.value and 
            prospecto.agente_asignado_id != user.id):
            return RedirectResponse(url="/prospectos?error=No tiene permisos para eliminar este prospecto", status_code=303)
        
        # ✅ SOFT DELETE: Marcar como eliminado en lugar de borrar
        prospecto.fecha_eliminacion = datetime.now()
        
        # Opcional: Cambiar estado a un estado especial
        if prospecto.estado not in [EstadoProspecto.GANADO.value, EstadoProspecto.CERRADO_PERDIDO.value]:
            prospecto.estado_anterior = prospecto.estado
            prospecto.estado = "eliminado"
        
        # Crear interacción de eliminación para trazabilidad
        interaccion = models.Interaccion(
            prospecto_id=prospecto_id,
            usuario_id=user.id,
            tipo_interaccion="sistema",
            descripcion=f"🗑️ Prospecto marcado como eliminado por {user.username}",
            estado_anterior=prospecto.estado_anterior or prospecto.estado,
            estado_nuevo="eliminado"
        )
        db.add(interaccion)
        
        db.commit()
        
        return RedirectResponse(url="/prospectos?success=Prospecto eliminado correctamente", status_code=303)
    
    except Exception as e:
        db.rollback()
        print(f"❌ Error eliminando prospecto: {e}")
        return RedirectResponse(url="/prospectos?error=Error al eliminar prospecto", status_code=303)

@app.post("/prospectos/{prospecto_id}/asignar")
async def asignar_agente(
    request: Request,
    prospecto_id: int,
    agente_id: int = Form(None),  # ID del agente a asignar (0 para desasignar)
    # ✅ PARÁMETROS PARA MANTENER FILTROS
    destino: str = Form(None),
    telefono: str = Form(None),
    medio_ingreso_id: str = Form(None),
    estado: str = Form(None),
    busqueda_global: str = Form(None),
    agente_filtro_id: str = Form(None),  # ✅ CAMBIADO: Filtro de agente (no confundir con asignación)
    # ✅ PARÁMETROS DE FECHA PARA FILTROS DESDE DASHBOARD
    fecha_inicio: str = Form(None),
    fecha_fin: str = Form(None),
    periodo: str = Form(None),
    tipo_filtro: str = Form(None),  # Para filtros desde dashboard
    valor_filtro: str = Form(None),  # Para filtros desde dashboard
    pagina: str = Form("1"),  # Para mantener paginación
    db: Session = Depends(database.get_db)
):
    user = await get_current_user(request, db)
    if not user or user.tipo_usuario not in [TipoUsuario.ADMINISTRADOR.value, TipoUsuario.SUPERVISOR.value]:
        raise HTTPException(status_code=403, detail="No tiene permisos para esta acción")
    
    try:
        prospecto = db.query(models.Prospecto).filter(models.Prospecto.id == prospecto_id).first()
        if not prospecto:
            raise HTTPException(status_code=404, detail="Prospecto no encontrado")
        
        # Si agente_id es 0 o vacío, desasignar (establecer None)
        if not agente_id or agente_id == 0:
            prospecto.agente_asignado_id = None
            mensaje = "Prospecto desasignado correctamente"
        else:
            # Verificar que el agente exista
            agente = db.query(models.Usuario).filter(
                models.Usuario.id == agente_id,
                models.Usuario.tipo_usuario == TipoUsuario.AGENTE.value
            ).first()
            if not agente:
                raise HTTPException(status_code=404, detail="Agente no encontrado")
            
            prospecto.agente_asignado_id = agente_id
            mensaje = f"Agente {agente.username} asignado correctamente"
            
            # ✅ CREAR NOTIFICACIÓN DE ASIGNACIÓN
            notificacion = models.Notificacion(
                usuario_id=agente.id,
                prospecto_id=prospecto.id,
                tipo="asignacion",
                mensaje=f"Te han asignado un nuevo prospecto: {prospecto.nombre} {prospecto.apellido or ''}",
                email_enviado=False
            )
            db.add(notificacion)
            
            # ✅ ENVIAR EMAIL AL AGENTE
            if agente.email:
                asunto = "Nuevo Prospecto Asignado 🚀"
                cuerpo = f"Hola {agente.username},\n\nSe te ha asignado el prospecto {prospecto.nombre} {prospecto.apellido}.\n\nIngresa al sistema para gestionarlo."
                enviado = enviar_notificacion_email(agente.email, asunto, cuerpo)
                notificacion.email_enviado = enviado
        
        db.commit()
        
        # ✅ DETERMINAR A DÓNDE REDIRIGIR
        redirect_url = "/prospectos"  # Por defecto
        
        # Si viene de un filtro del dashboard, redirigir a esa vista
        if tipo_filtro and valor_filtro:
            redirect_url = "/prospectos/filtro"
        
        # ✅ CONSTRUIR PARÁMETROS DE CONSULTA
        params = []
        
        # Parámetros generales de filtros
        if destino:
            params.append(f"destino={destino}")
        if telefono:
            params.append(f"telefono={telefono}")
        if medio_ingreso_id and medio_ingreso_id != 'todos':
            params.append(f"medio_ingreso_id={medio_ingreso_id}")
        if estado and estado != 'todos':
            params.append(f"estado={estado}")
        if busqueda_global:
            params.append(f"busqueda_global={busqueda_global}")
        if agente_filtro_id and agente_filtro_id != 'todos':
            params.append(f"agente_asignado_id={agente_filtro_id}")
        
        # Parámetros específicos para filtros desde dashboard
        if tipo_filtro:
            params.append(f"tipo_filtro={tipo_filtro}")
        if valor_filtro:
            params.append(f"valor_filtro={valor_filtro}")
        if fecha_inicio:
            params.append(f"fecha_inicio={fecha_inicio}")
        if fecha_fin:
            params.append(f"fecha_fin={fecha_fin}")
        if periodo:
            params.append(f"periodo={periodo}")
        if pagina and pagina != "1":
            params.append(f"pagina={pagina}")
        
        # Agregar mensaje de éxito
        if params:
            redirect_url += "?" + "&".join(params) + f"&success={mensaje}"
        else:
            redirect_url += f"?success={mensaje}"
        
        return RedirectResponse(url=redirect_url, status_code=303)
    
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        print(f"❌ Error asignando agente: {e}")
        return RedirectResponse(url="/prospectos?error=Error al asignar agente", status_code=303)

# ========== GESTIÓN DE INTERACCIONES ==========

@app.get("/prospectos/{prospecto_id}/seguimiento")
async def ver_seguimiento(
    request: Request,
    prospecto_id: int,
    db: Session = Depends(database.get_db)
):
    user = await get_current_user(request, db)
    if not user:
        return RedirectResponse(url="/", status_code=303)
    
    prospecto = db.query(models.Prospecto).filter(models.Prospecto.id == prospecto_id).first()
    if not prospecto:
        return RedirectResponse(url="/prospectos?error=Prospecto no encontrado", status_code=303)
    
    # Verificar permisos
    if (user.tipo_usuario == TipoUsuario.AGENTE.value and 
        prospecto.agente_asignado_id != user.id):
        return RedirectResponse(url="/prospectos?error=No tiene permisos para ver este prospecto", status_code=303)
    
    return templates.TemplateResponse("seguimiento_prospecto.html", {
        "request": request,
        "prospecto": prospecto,
        "current_user": user,
        "estados_prospecto": [estado.value for estado in EstadoProspecto]
    })

@app.post("/prospectos/{prospecto_id}/interaccion")
async def registrar_interaccion(
    request: Request,
    prospecto_id: int,
    descripcion: str = Form(...),
    tipo_interaccion: str = Form("general"),
    cambio_estado: str = Form(None),
    fecha_proximo_contacto: str = Form(None),
    db: Session = Depends(database.get_db)
):
    user = await get_current_user(request, db)
    if not user:
        return RedirectResponse(url="/", status_code=303)
    
    try:
        prospecto = db.query(models.Prospecto).filter(models.Prospecto.id == prospecto_id).first()
        if not prospecto:
            return RedirectResponse(url="/prospectos?error=Prospecto no encontrado", status_code=303)
        
        # Verificar permisos
        if (user.tipo_usuario == TipoUsuario.AGENTE.value and 
            prospecto.agente_asignado_id != user.id):
            return RedirectResponse(url="/prospectos?error=No tiene permisos para este prospecto", status_code=303)
        
        # ✅ VALIDAR TRANSICIÓN DE ESTADOS (NO PERMITIR REGRESAR)
        if cambio_estado and prospecto.estado:
            estados_orden = [
                EstadoProspecto.NUEVO.value,
                EstadoProspecto.EN_SEGUIMIENTO.value, 
                EstadoProspecto.COTIZADO.value,
                # Estados finales (mismo nivel)
                EstadoProspecto.GANADO.value,
                EstadoProspecto.CERRADO_PERDIDO.value
            ]
            
            estado_actual_idx = estados_orden.index(prospecto.estado) if prospecto.estado in estados_orden else -1
            estado_nuevo_idx = estados_orden.index(cambio_estado) if cambio_estado in estados_orden else -1
            
            # No permitir regresar a estados anteriores (excepto reactivación por admin/supervisor)
            if (estado_nuevo_idx < estado_actual_idx and 
                estado_actual_idx >= 2 and  # Solo validar desde COTIZADO hacia atrás
                user.tipo_usuario not in [TipoUsuario.ADMINISTRADOR.value, TipoUsuario.SUPERVISOR.value]):
                return RedirectResponse(
                    url=f"/prospectos/{prospecto_id}/seguimiento?error=No puede regresar a un estado anterior", 
                    status_code=303
                )
        
        # Validar cambio de estado a CERRADO_PERDIDO
        if cambio_estado == EstadoProspecto.CERRADO_PERDIDO.value and not descripcion.strip():
            return RedirectResponse(
                url=f"/prospectos/{prospecto_id}/seguimiento?error=Debe agregar un comentario al cerrar el prospecto", 
                status_code=303
            )
        
        # Registrar historial de estado ANTES del cambio
        if cambio_estado and cambio_estado != prospecto.estado:
            historial = models.HistorialEstado(
                prospecto_id=prospecto_id,
                estado_anterior=prospecto.estado,
                estado_nuevo=cambio_estado,
                usuario_id=user.id,
                comentario=descripcion
            )
            db.add(historial)
        
        # Registrar interacción
        estado_anterior = prospecto.estado
        interaccion = models.Interaccion(
            prospecto_id=prospecto_id,
            usuario_id=user.id,
            tipo_interaccion=tipo_interaccion,
            descripcion=descripcion,
            estado_anterior=estado_anterior,
            estado_nuevo=cambio_estado
        )
        
        db.add(interaccion)
        
        # ✅ CREAR NOTIFICACIÓN DE SEGUIMIENTO (SI SE PROGRAMA)
        if fecha_proximo_contacto:
            try:
                # El formato de input datetime-local es "YYYY-MM-DDTHH:MM"
                fecha_prog = datetime.strptime(fecha_proximo_contacto, "%Y-%m-%dT%H:%M")
                
                notificacion = models.Notificacion(
                    usuario_id=user.id,
                    prospecto_id=prospecto_id,
                    tipo="seguimiento",
                    mensaje=f"Recordatorio: {descripcion[:50]}...",
                    fecha_programada=fecha_prog,
                    email_enviado=False
                )
                db.add(notificacion)
                
                # Feedback visual en el log/email
                if user.email:
                    enviar_notificacion_email(
                        user.email, 
                        "Recordatorio Programado 📅", 
                        f"Has programado un seguimiento para el prospecto {prospecto.nombre} el {fecha_prog}."
                    )
            except ValueError:
                print(f"❌ Error formato fecha recordatorio: {fecha_proximo_contacto}")
        
        # ✅ REGISTRAR ESTADÍSTICA DE COTIZACIÓN (SIEMPRE CREAR NUEVA)
        if (cambio_estado == EstadoProspecto.COTIZADO.value and 
            estado_anterior != EstadoProspecto.COTIZADO.value and
            prospecto.agente_asignado_id):
            
            # ✅ SIEMPRE CREAR NUEVA ESTADÍSTICA (permitir múltiples cotizaciones)
            estadistica = models.EstadisticaCotizacion(
                agente_id=prospecto.agente_asignado_id,
                prospecto_id=prospecto_id,
                fecha_cotizacion=datetime.now().date()
            )
            db.add(estadistica)
            
            # ✅ GENERAR ID DE COTIZACIÓN ÚNICO
            db.flush()
            estadistica.generar_id_cotizacion()
            
            # ✅ ASIGNAR ID DE COTIZACIÓN AL PROSPECTO (última cotización)
            prospecto.id_cotizacion = estadistica.id_cotizacion
            print(f"✅ Nueva cotización generada al cambiar estado: {estadistica.id_cotizacion}")

        
        # Actualizar estado del prospecto si hay cambio
        if cambio_estado:
            prospecto.estado = cambio_estado
            
            # ✅ NUEVO: Auto-poblar fecha_compra cuando se marca como GANADO
            if cambio_estado == EstadoProspecto.GANADO.value and not prospecto.fecha_compra:
                prospecto.fecha_compra = datetime.now().date()
        
        db.commit()
        
        return RedirectResponse(
            url=f"/prospectos/{prospecto_id}/seguimiento?success=Interacción registrada", 
            status_code=303
        )
    
    except Exception as e:
        db.rollback()
        print(f"❌ Error registrando interacción: {e}")
        return RedirectResponse(
            url=f"/prospectos/{prospecto_id}/seguimiento?error=Error al registrar interacción", 
            status_code=303
        )
    

# ========== GESTIÓN DE DOCUMENTOS ==========

@app.post("/prospectos/{prospecto_id}/documento")
async def subir_documento(
    request: Request,
    prospecto_id: int,
    archivo: UploadFile = File(...),
    tipo_documento: str = Form("cotizacion"),
    descripcion: str = Form(None),
    db: Session = Depends(database.get_db)
):
    user = await get_current_user(request, db)
    if not user:
        return RedirectResponse(url="/", status_code=303)
    
    try:
        prospecto = db.query(models.Prospecto).filter(models.Prospecto.id == prospecto_id).first()
        if not prospecto:
            return RedirectResponse(url="/prospectos?error=Prospecto no encontrado", status_code=303)
        
        # Verificar permisos
        if (user.tipo_usuario == TipoUsuario.AGENTE.value and 
            prospecto.agente_asignado_id != user.id):
            return RedirectResponse(url="/prospectos?error=No tiene permisos para este prospecto", status_code=303)
        
        # Validar tipo de archivo
        allowed_extensions = {'.pdf', '.doc', '.docx', '.xls', '.xlsx', '.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp'}
        file_ext = os.path.splitext(archivo.filename)[1].lower()
        
        if file_ext not in allowed_extensions:
            return RedirectResponse(
                url=f"/prospectos/{prospecto_id}/seguimiento?error=Solo se permiten archivos PDF, Office e imágenes",
                status_code=303
            )

        # ✅ LEER Y COMPRIMIR EL ARCHIVO
        file_content = await archivo.read()
        from utils.compression import procesar_archivo_para_subida
        
        print(f"Tamaño original de {archivo.filename}: {len(file_content)} bytes")
        file_content_compressed = procesar_archivo_para_subida(file_content, archivo.filename)
        print(f"Tamaño comprimido de {archivo.filename}: {len(file_content_compressed)} bytes")
        
        # ✅ SUBIR A GOOGLE DRIVE
        from services.google_drive import upload_to_drive
        folder_id = os.getenv("GOOGLE_DRIVE_FOLDER_ID")
        
        # Obtener mimetype básico
        mimetype_map = {
            '.pdf': 'application/pdf', '.jpg': 'image/jpeg', '.jpeg': 'image/jpeg',
            '.png': 'image/png', '.docx': 'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
            '.xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        }
        mime_type = mimetype_map.get(file_ext, 'application/octet-stream')
        
        timestamp_name = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{archivo.filename}"
        
        # Subir a Drive. Devuelve el ID del archivo (ej: '1B2a_xY...')
        drive_file_id = upload_to_drive(file_content_compressed, timestamp_name, mime_type, folder_id)
        
        if not drive_file_id:
            # Fallback en caso de que Google Drive no esté configurado o falle, guardamos localmente
            print("⚠️ Falla al subir a Drive. Guardando localmente como fallback.")
            ruta_fecha = obtener_ruta_upload_por_fecha()
            ruta_archivo_completa = os.path.join(ruta_fecha, timestamp_name)
            with open(ruta_archivo_completa, "wb") as buffer:
                buffer.write(file_content_compressed)
            ruta_guardado = os.path.relpath(ruta_archivo_completa, UPLOAD_DIR)
        else:
            # Si se subió con éxito a Drive, guardamos el ID (con un prefijo para identificarlo luego)
            ruta_guardado = f"gdrive://{drive_file_id}"
        
        # ✅ REGISTRAR DOCUMENTO EN BD
        documento = models.Documento(
            prospecto_id=prospecto_id,
            usuario_id=user.id,
            nombre_archivo=archivo.filename,
            tipo_documento=tipo_documento,
            ruta_archivo=ruta_guardado,  # ID de Drive o ruta local
            descripcion=descripcion
        )
        
        db.add(documento)
        db.flush()  # Para obtener el ID antes del commit
        
        # ✅ GENERAR ID DE DOCUMENTO ÚNICO
        documento.generar_id_documento()
        
        # ✅ CAMBIAR ESTADO A COTIZADO SI SE SUBE UNA COTIZACIÓN Y NO ESTÁ EN ESTADOS FINALES
        if tipo_documento == "cotizacion":
            estado_anterior = prospecto.estado
            
            # No sobreescribir el estado si ya está ganado, perdido o cancelado
            estados_protegidos = [
                EstadoProspecto.GANADO.value, 
                EstadoProspecto.CERRADO_PERDIDO.value,
                EstadoProspecto.VENTA_CANCELADA.value
            ]
            
            if prospecto.estado not in estados_protegidos:
                prospecto.estado = EstadoProspecto.COTIZADO.value
            
            # ✅ REGISTRAR NUEVA ESTADÍSTICA DE COTIZACIÓN (SIEMPRE CREAR NUEVA)
            # Permitir múltiples cotizaciones por solicitud
            estadistica = models.EstadisticaCotizacion(
                agente_id=prospecto.agente_asignado_id or user.id,
                prospecto_id=prospecto_id,
                fecha_cotizacion=datetime.now().date()
            )
            db.add(estadistica)
            db.flush()  # Para obtener el ID
            
            # ✅ GENERAR ID DE COTIZACIÓN ÚNICO
            estadistica.generar_id_cotizacion()
            
            # ✅ ASIGNAR ID DE COTIZACIÓN AL PROSPECTO (última cotización)
            prospecto.id_cotizacion = estadistica.id_cotizacion
            print(f"✅ Nueva cotización generada: {estadistica.id_cotizacion}")

            
            # Registrar interacción automática de cambio de estado o subida
            if prospecto.estado != estado_anterior:
                descripcion = f"Se subió cotización y cambió estado a Cotizado: {archivo.filename}"
            else:
                descripcion = f"Se subió cotización: {archivo.filename} (mantiene estado {estado_anterior})"
                
            interaccion = models.Interaccion(
                prospecto_id=prospecto_id,
                usuario_id=user.id,
                tipo_interaccion="documento",
                descripcion=descripcion,
                estado_anterior=estado_anterior,
                estado_nuevo=prospecto.estado
            )
            db.add(interaccion)
        
        # Registrar interacción para el documento
        interaccion_doc = models.Interaccion(
            prospecto_id=prospecto_id,
            usuario_id=user.id,
            tipo_interaccion="documento",
            descripcion=f"Documento subido: {archivo.filename} ({tipo_documento})",
            estado_anterior=prospecto.estado,
            estado_nuevo=prospecto.estado
        )
        db.add(interaccion_doc)
        
        db.commit()
        
        return RedirectResponse(
            url=f"/prospectos/{prospecto_id}/seguimiento?success=Documento subido correctamente", 
            status_code=303
        )
    
    except Exception as e:
        db.rollback()
        print(f"❌ Error subiendo documento: {e}")
        import traceback
        traceback.print_exc()
        return RedirectResponse(
            url=f"/prospectos/{prospecto_id}/seguimiento?error=Error al subir documento", 
            status_code=303
        )

# ✅ NUEVO ENDPOINT: Descargar documento desde Google Drive
@app.get("/documentos/{documento_id}/descargar")
async def descargar_documento(
    request: Request,
    documento_id: int,
    db: Session = Depends(database.get_db)
):
    user = await get_current_user(request, db)
    if not user:
        return RedirectResponse(url="/", status_code=303)
        
    documento = db.query(models.Documento).filter(models.Documento.id == documento_id).first()
    if not documento:
        return RedirectResponse(url="/prospectos?error=Documento no encontrado", status_code=303)
        
    # Verificar si es de drive
    if documento.ruta_archivo and documento.ruta_archivo.startswith("gdrive://"):
        file_id = documento.ruta_archivo.replace("gdrive://", "")
        from services.google_drive import get_file_url
        url = get_file_url(file_id)
        # Redirigimos al visor web de Google Drive
        return RedirectResponse(url=url)
    
    # Si es local, redirigimos a la ruta relativa que FastApi sirve
    # ya que FastAPI está sirviendo el folder entero en /uploads,
    # y ruta_archivo es por ejemplo '123/archivo.pdf', podemos redirigir a /uploads/123/archivo.pdf
    return RedirectResponse(url=f"/uploads/{documento.ruta_archivo}")

# ✅ NUEVO ENDPOINT: Búsqueda por ID
@app.get("/busqueda_ids", response_class=HTMLResponse)
async def buscar_por_id(
    request: Request,
    tipo_id: str = Query("cliente"),  # cliente, solicitud, cotizacion, documento
    valor_id: str = Query(None),
    db: Session = Depends(database.get_db)
):
    user = await get_current_user(request, db)
    if not user:
        return RedirectResponse(url="/", status_code=303)
    
    resultados = []
    tipo_busqueda = ""
    
    if valor_id:
        valor_id = valor_id.upper().strip()
        
        if tipo_id == "cliente":
            # Buscar por ID de cliente (puede traer múltiples solicitudes)
            resultados = db.query(models.Prospecto).filter(
                models.Prospecto.id_cliente.ilike(f"%{valor_id}%")
            ).all()
            tipo_busqueda = f"Clientes con ID: {valor_id}"
        
        elif tipo_id == "solicitud":
            # ✅ NUEVO: Buscar por ID de solicitud (único)
            resultados = db.query(models.Prospecto).filter(
                models.Prospecto.id_solicitud.ilike(f"%{valor_id}%")
            ).all()
            tipo_busqueda = f"Solicitudes con ID: {valor_id}"
            
        elif tipo_id == "cotizacion":
            # Buscar por ID de cotización
            estadisticas = db.query(models.EstadisticaCotizacion).filter(
                models.EstadisticaCotizacion.id_cotizacion.ilike(f"%{valor_id}%")
            ).all()
            # Obtener prospectos relacionados
            for stats in estadisticas:
                prospecto = db.query(models.Prospecto).filter(
                    models.Prospecto.id == stats.prospecto_id
                ).first()
                if prospecto:
                    resultados.append(prospecto)
            tipo_busqueda = f"Cotizaciones con ID: {valor_id}"
            
        elif tipo_id == "documento":
            # Buscar por ID de documento
            documentos = db.query(models.Documento).filter(
                models.Documento.id_documento.ilike(f"%{valor_id}%")
            ).all()
            # Obtener prospectos relacionados
            for doc in documentos:
                prospecto = db.query(models.Prospecto).filter(
                    models.Prospecto.id == doc.prospecto_id
                ).first()
                if prospecto:
                    resultados.append(prospecto)
            tipo_busqueda = f"Documentos con ID: {valor_id}"
    
    return templates.TemplateResponse("busqueda_ids.html", {
        "request": request,
        "current_user": user,
        "resultados": resultados,
        "tipo_busqueda": tipo_busqueda,
        "tipo_id_activo": tipo_id,
        "valor_id_buscado": valor_id
    })




# ========== GESTIÓN DE USUARIOS (SOLO ADMIN) ==========

@app.get("/usuarios", response_class=HTMLResponse)
async def listar_usuarios(
    request: Request,
    page: int = Query(1, ge=1),
    limit: int = Query(20, ge=1, le=100),
    filtro_estado: str = Query("activos"),  # activos, inactivos, todos
    db: Session = Depends(database.get_db),
    user: models.Usuario = Depends(require_admin)
):
    """Lista usuarios con paginación y filtro de estado"""
    
    # Query base
    query = db.query(models.Usuario)
    
    # Aplicar filtro de estado
    if filtro_estado == "activos":
        query = query.filter(models.Usuario.activo == 1)
    elif filtro_estado == "inactivos":
        query = query.filter(models.Usuario.activo == 0)
    # Si es "todos", no filtrar
    
    # Contar total
    total_usuarios = query.count()
    
    # Calcular paginación
    total_pages = (total_usuarios + limit - 1) // limit
    offset = (page - 1) * limit
    
    # Obtener usuarios paginados
    usuarios = query.order_by(models.Usuario.fecha_creacion.desc()).offset(offset).limit(limit).all()
    
    return templates.TemplateResponse("usuarios/lista.html", {
        "request": request,
        "current_user": user,
        "usuarios": usuarios,
        "tipos_usuario": [tipo.value for tipo in TipoUsuario],
        "page": page,
        "limit": limit,
        "total_pages": total_pages,
        "total_usuarios": total_usuarios,
        "filtro_estado": filtro_estado
    })

@app.post("/usuarios")
async def crear_usuario(
    request: Request,
    username: str = Form(...),
    email: str = Form(...),
    password: str = Form(...),
    tipo_usuario: str = Form(...),
    db: Session = Depends(database.get_db),
    user: models.Usuario = Depends(require_admin)
):
    try:
        # Verificar si usuario ya existe
        existing_user = db.query(models.Usuario).filter(
            (models.Usuario.username == username) | (models.Usuario.email == email)
        ).first()
        
        if existing_user:
            return RedirectResponse(url="/usuarios?error=Usuario o email ya existen", status_code=303)
        
        # Crear usuario
        nuevo_usuario = models.Usuario(
            username=username,
            email=email,
            hashed_password=auth.get_password_hash(password),
            tipo_usuario=tipo_usuario
        )
        
        db.add(nuevo_usuario)
        db.commit()
        
        return RedirectResponse(url="/usuarios?success=Usuario creado correctamente", status_code=303)
    
    except Exception as e:
        db.rollback()
        print(f"❌ Error creating user: {e}")
        return RedirectResponse(url="/usuarios?error=Error al crear usuario", status_code=303)

@app.post("/usuarios/{usuario_id}/editar")
async def editar_usuario(
    request: Request,
    usuario_id: int,
    username: str = Form(...),
    email: str = Form(...),
    tipo_usuario: str = Form(...),
    password: str = Form(None),
    db: Session = Depends(database.get_db),
    user: models.Usuario = Depends(require_admin)
):
    try:
        usuario = db.query(models.Usuario).filter(models.Usuario.id == usuario_id).first()
        if not usuario:
            return RedirectResponse(url="/usuarios?error=Usuario no encontrado", status_code=303)
        
        # Verificar si username/email ya existen en otros usuarios
        existing_user = db.query(models.Usuario).filter(
            (models.Usuario.username == username) | (models.Usuario.email == email)
        ).filter(models.Usuario.id != usuario_id).first()
        
        if existing_user:
            return RedirectResponse(url="/usuarios?error=Usuario o email ya existen", status_code=303)
        
        # Actualizar datos
        usuario.username = username
        usuario.email = email
        usuario.tipo_usuario = tipo_usuario
        
        # Actualizar contraseña si se proporcionó
        if password:
            usuario.hashed_password = auth.get_password_hash(password)
        
        db.commit()
        
        return RedirectResponse(url="/usuarios?success=Usuario actualizado correctamente", status_code=303)
    
    except Exception as e:
        db.rollback()
        print(f"❌ Error updating user: {e}")
        return RedirectResponse(url="/usuarios?error=Error al actualizar usuario", status_code=303)

@app.post("/usuarios/{usuario_id}/eliminar")
async def eliminar_usuario(
    request: Request,
    usuario_id: int,
    db: Session = Depends(database.get_db),
    user: models.Usuario = Depends(require_admin)
):
    try:
        # No permitir eliminar el propio usuario
        if usuario_id == user.id:
            return RedirectResponse(url="/usuarios?error=No puede eliminar su propio usuario", status_code=303)
        
        usuario = db.query(models.Usuario).filter(models.Usuario.id == usuario_id).first()
        if not usuario:
            return RedirectResponse(url="/usuarios?error=Usuario no encontrado", status_code=303)
        
        db.delete(usuario)
        db.commit()
        
        return RedirectResponse(url="/usuarios?success=Usuario eliminado correctamente", status_code=303)
    
    except Exception as e:
        db.rollback()
        print(f"❌ Error deleting user: {e}")
        return RedirectResponse(url="/usuarios?error=Error al eliminar usuario", status_code=303)

@app.post("/usuarios/{usuario_id}/desactivar")
async def desactivar_usuario(
    request: Request,
    usuario_id: int,
    db: Session = Depends(database.get_db),
    user: models.Usuario = Depends(require_admin)
):
    """Desactiva un usuario y reasigna sus prospectos activos a servicio_cliente"""
    try:
        # No permitir desactivar el propio usuario
        if usuario_id == user.id:
            return RedirectResponse(url="/usuarios?error=No puede desactivar su propio usuario", status_code=303)
        
        usuario = db.query(models.Usuario).filter(models.Usuario.id == usuario_id).first()
        if not usuario:
            return RedirectResponse(url="/usuarios?error=Usuario no encontrado", status_code=303)
        
        if usuario.activo == 0:
            return RedirectResponse(url="/usuarios?error=El usuario ya está inactivo", status_code=303)
        
        # Buscar usuario servicio_cliente
        servicio_cliente = db.query(models.Usuario).filter(
            models.Usuario.username == "servicio_cliente"
        ).first()
        
        if not servicio_cliente:
            return RedirectResponse(url="/usuarios?error=Usuario servicio_cliente no encontrado", status_code=303)
        
        # Reasignar prospectos activos
        prospectos_activos = db.query(models.Prospecto).filter(
            models.Prospecto.agente_asignado_id == usuario_id,
            models.Prospecto.estado.in_([
                EstadoProspecto.NUEVO.value,
                EstadoProspecto.EN_SEGUIMIENTO.value,
                EstadoProspecto.COTIZADO.value
            ])
        ).all()
        
        prospectos_reasignados = 0
        for prospecto in prospectos_activos:
            prospecto.agente_original_id = usuario_id
            prospecto.agente_asignado_id = servicio_cliente.id
            prospectos_reasignados += 1
        
        # Desactivar usuario
        usuario.activo = 0
        usuario.email = "servicioclientetravelhouse@gmail.com"
        
        db.commit()
        
        mensaje = f"Usuario desactivado correctamente. {prospectos_reasignados} prospectos reasignados a servicio_cliente"
        return RedirectResponse(url=f"/usuarios?success={mensaje}", status_code=303)
    
    except Exception as e:
        db.rollback()
        print(f"Error desactivando usuario: {e}")
        return RedirectResponse(url="/usuarios?error=Error al desactivar usuario", status_code=303)

@app.post("/usuarios/{usuario_id}/reactivar")
async def reactivar_usuario(
    request: Request,
    usuario_id: int,
    email: str = Form(...),
    db: Session = Depends(database.get_db),
    user: models.Usuario = Depends(require_admin)
):
    """Reactiva un usuario inactivo"""
    try:
        usuario = db.query(models.Usuario).filter(models.Usuario.id == usuario_id).first()
        if not usuario:
            return RedirectResponse(url="/usuarios?error=Usuario no encontrado", status_code=303)
        
        if usuario.activo == 1:
            return RedirectResponse(url="/usuarios?error=El usuario ya está activo", status_code=303)
        
        # Validar que el email no esté duplicado entre usuarios activos
        email_existente = db.query(models.Usuario).filter(
            models.Usuario.email == email,
            models.Usuario.activo == 1,
            models.Usuario.id != usuario_id
        ).first()
        
        if email_existente:
            return RedirectResponse(url="/usuarios?error=El email ya está en uso por otro usuario activo", status_code=303)
        
        # Reactivar usuario
        usuario.activo = 1
        usuario.email = email
        
        db.commit()
        
        return RedirectResponse(url="/usuarios?success=Usuario reactivado correctamente", status_code=303)
    
    except Exception as e:
        db.rollback()
        print(f"Error reactivando usuario: {e}")
        return RedirectResponse(url="/usuarios?error=Error al reactivar usuario", status_code=303)

# ========== HISTORIAL DE PROSPECTOS CERRADOS ==========

@app.get("/prospectos/cerrados", response_class=HTMLResponse)
async def listar_prospectos_cerrados(
    request: Request,
    fecha_registro_desde: str = Query(None),
    fecha_registro_hasta: str = Query(None),
    fecha_cierre_desde: str = Query(None),
    fecha_cierre_hasta: str = Query(None),
    destino: str = Query(None),
    agente_asignado_id: str = Query(None),
    page: int = Query(1, ge=1),
    limit: int = Query(10, ge=1, le=100),
    db: Session = Depends(database.get_db)
):
    user = await get_current_user(request, db)
    
    if not user:
        return RedirectResponse(url="/", status_code=303)
    
    # Construir query para prospectos cerrados/ganados
    query = db.query(models.Prospecto).filter(
        models.Prospecto.estado.in_([EstadoProspecto.CERRADO_PERDIDO.value, EstadoProspecto.GANADO.value])
    )
    
    # Aplicar filtros según rol
    if user.tipo_usuario == TipoUsuario.AGENTE.value:
        query = query.filter(models.Prospecto.agente_asignado_id == user.id)
    
    # Filtros de fechas de registro
    if fecha_registro_desde:
        try:
            fecha_desde = datetime.strptime(fecha_registro_desde, "%d/%m/%Y").date()
            query = query.filter(models.Prospecto.fecha_registro >= fecha_desde)
        except ValueError:
            pass
    
    if fecha_registro_hasta:
        try:
            fecha_hasta = datetime.strptime(fecha_registro_hasta, "%d/%m/%Y").date()
            query = query.filter(models.Prospecto.fecha_registro <= fecha_hasta)
        except ValueError:
            pass
    
    # Filtros de fechas de cierre (última interacción con cambio de estado)
    if fecha_cierre_desde or fecha_cierre_hasta:
        subquery = db.query(models.Interaccion.prospecto_id).filter(
            models.Interaccion.estado_nuevo.in_([EstadoProspecto.CERRADO_PERDIDO.value, EstadoProspecto.GANADO.value])
        )
        
        if fecha_cierre_desde:
            try:
                fecha_cierre_desde_date = datetime.strptime(fecha_cierre_desde, "%d/%m/%Y").date()
                subquery = subquery.filter(models.Interaccion.fecha_creacion >= fecha_cierre_desde_date)
            except ValueError:
                pass
        
        if fecha_cierre_hasta:
            try:
                fecha_cierre_hasta_date = datetime.strptime(fecha_cierre_hasta, "%d/%m/%Y").date()
                subquery = subquery.filter(models.Interaccion.fecha_creacion <= fecha_cierre_hasta_date)
            except ValueError:
                pass
        
        query = query.filter(models.Prospecto.id.in_(subquery))
    
    # Otros filtros
    if destino:
        # ✅ BÚSQUEDA GENERAL EN CERRADOS (Nombre, Email, Teléfono, Destino)
        search_term = f"%{destino}%"
        query = query.filter(or_(
            models.Prospecto.destino.ilike(search_term),
            models.Prospecto.nombre.ilike(search_term),
            models.Prospecto.apellido.ilike(search_term),
            models.Prospecto.correo_electronico.ilike(search_term),
            models.Prospecto.telefono.ilike(search_term)
        ))
    
    if agente_asignado_id and agente_asignado_id != "todos" and user.tipo_usuario in [TipoUsuario.ADMINISTRADOR.value, TipoUsuario.SUPERVISOR.value]:
        query = query.filter(models.Prospecto.agente_asignado_id == int(agente_asignado_id))
    
    # ✅ PAGINACIÓN
    query = query.order_by(models.Prospecto.fecha_registro.desc())
    
    total_registros = query.count()
    total_pages = (total_registros + limit - 1) // limit
    
    # Asegurar que la página solicitada sea válida
    if page > total_pages and total_pages > 0:
        page = total_pages
    if page < 1:
        page = 1
        
    offset = (page - 1) * limit
    prospectos_cerrados = query.offset(offset).limit(limit).all()
    
    # Obtener datos para filtros
    agentes = db.query(models.Usuario).filter(
        models.Usuario.tipo_usuario == TipoUsuario.AGENTE.value
    ).all()
    
    return templates.TemplateResponse("prospectos_cerrados.html", {
        "request": request,
        "prospectos": prospectos_cerrados,
        "current_user": user,
        "agentes": agentes,
        "filtros_activos": {
            "fecha_registro_desde": fecha_registro_desde,
            "fecha_registro_hasta": fecha_registro_hasta,
            "fecha_cierre_desde": fecha_cierre_desde,
            "fecha_cierre_hasta": fecha_cierre_hasta,
            "destino": destino,
            "agente_asignado_id": agente_asignado_id
        },
        # Info Paginación
        "page": page,
        "limit": limit,
        "total_pages": total_pages,
        "total_registros": total_registros
    })

@app.post("/prospectos/{prospecto_id}/reactivar")
async def reactivar_prospecto(
    request: Request,
    prospecto_id: int,
    db: Session = Depends(database.get_db)
):
    user = await get_current_user(request, db)
    if not user:
        return RedirectResponse(url="/", status_code=303)
    
    try:
        prospecto = db.query(models.Prospecto).filter(models.Prospecto.id == prospecto_id).first()
        if not prospecto:
            return RedirectResponse(url="/prospectos/cerrados?error=Prospecto no encontrado", status_code=303)
        
        # Verificar permisos
        if (user.tipo_usuario == TipoUsuario.AGENTE.value and 
            prospecto.agente_asignado_id != user.id):
            return RedirectResponse(url="/prospectos/cerrados?error=No tiene permisos para reactivar este prospecto", status_code=303)
        
        # Reactivar prospecto
        estado_anterior = prospecto.estado
        prospecto.estado = EstadoProspecto.EN_SEGUIMIENTO.value
        
        # Registrar interacción de reactivación
        interaccion = models.Interaccion(
            prospecto_id=prospecto_id,
            usuario_id=user.id,
            tipo_interaccion="sistema",
            descripcion=f"Prospecto reactivado desde estado: {estado_anterior}",
            estado_anterior=estado_anterior,
            estado_nuevo=EstadoProspecto.EN_SEGUIMIENTO.value
        )
        
        db.add(interaccion)
        db.commit()
        
        return RedirectResponse(url="/prospectos/cerrados?success=Prospecto reactivado correctamente", status_code=303)
    
    except Exception as e:
        db.rollback()
        print(f"❌ Error reactivando prospecto: {e}")
        return RedirectResponse(url="/prospectos/cerrados?error=Error al reactivar prospecto", status_code=303)


# Endpoint para verificar autenticación
@app.get("/check-auth")
async def check_auth(request: Request, db: Session = Depends(database.get_db)):
    user = await get_current_user(request, db)
    if user:
        return {
            "authenticated": True, 
            "user": user.username,
            "user_type": user.tipo_usuario,
            "active_sessions": len(active_sessions)
        }
    else:
        return {
            "authenticated": False, 
            "active_sessions": len(active_sessions)
        }

#Exportar a Excel
@app.get("/prospectos/exportar/excel")
async def exportar_prospectos_excel(
    request: Request,
    db: Session = Depends(database.get_db),
    user: models.Usuario = Depends(get_current_user)
):
    try:
        # Obtener prospectos según permisos
        if user.tipo_usuario == TipoUsuario.AGENTE.value:
            prospectos = db.query(models.Prospecto).filter(
                models.Prospecto.agente_asignado_id == user.id
            ).all()
        else:
            prospectos = db.query(models.Prospecto).all()
        
        # Convertir a DataFrame
        data = []
        for p in prospectos:
            data.append({
                'ID': p.id,
                'Nombre': f"{p.nombre or ''} {p.apellido or ''}",
                'Email': p.correo_electronico or '',
                'Teléfono': p.telefono or '',
                'Destino': p.destino or '',
                'Estado': p.estado,
                'Agente': p.agente_asignado.username if p.agente_asignado else 'Sin asignar',
                'Fecha Registro': p.fecha_registro.strftime('%d/%m/%Y'),
                'Medio Ingreso': p.medio_ingreso.nombre
            })
        
        df = pd.DataFrame(data)
        
        # Crear Excel en memoria
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Prospectos', index=False)
        
        output.seek(0)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=prospectos.xlsx"}
        )
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error exportando: {str(e)}")

# ✅ PANEL DE HISTORIAL DE CLIENTE MEJORADO
@app.get("/clientes/historial", response_class=HTMLResponse)
async def historial_cliente(
    request: Request,
    busqueda: str = Query(None),
    telefono: str = Query(None),
    fecha_busqueda: str = Query(None),
    db: Session = Depends(database.get_db)
):
    user = await get_current_user(request, db)
    if not user:
        return RedirectResponse(url="/", status_code=303)
    
    cliente_principal = None
    prospectos = []
    
    query = db.query(models.Prospecto)
    
    # ✅ LÓGICA DE BÚSQUEDA AVANZADA
    filtros = []
    
    # 1. Búsqueda por término (Teléfono, Email, Nombre)
    if busqueda:
        term = f"%{busqueda}%"
        filtros.append(or_(
            models.Prospecto.telefono.ilike(term),
            models.Prospecto.telefono_secundario.ilike(term),
            models.Prospecto.correo_electronico.ilike(term),
            models.Prospecto.nombre.ilike(term),
            models.Prospecto.apellido.ilike(term)
        ))
    
    # 2. Búsqueda por teléfono específico (compatibilidad anterior)
    if telefono:
        filtros.append(or_(
            models.Prospecto.telefono == telefono,
            models.Prospecto.telefono_secundario == telefono
        ))
        
    # 3. Búsqueda por fecha exacta
    if fecha_busqueda:
        try:
            fecha_dt = datetime.strptime(fecha_busqueda, "%d/%m/%Y").date()
            # FIltrar por rango del día completo
            fecha_inicio = datetime.combine(fecha_dt, datetime.min.time())
            fecha_fin = datetime.combine(fecha_dt, datetime.max.time())
            filtros.append(and_(
                models.Prospecto.fecha_registro >= fecha_inicio,
                models.Prospecto.fecha_registro <= fecha_fin
            ))
        except ValueError:
            pass

    if filtros:
        query = query.filter(and_(*filtros))
        prospectos = query.order_by(models.Prospecto.fecha_registro.desc()).all()
        
        if prospectos:
            cliente_principal = prospectos[0]
            print(f"🔍 Encontrados {len(prospectos)} registros en historial")
    
    return templates.TemplateResponse("historial_cliente.html", {
        "request": request,
        "current_user": user,
        "cliente": cliente_principal,
        "prospectos": prospectos,
        "busqueda_activa": busqueda or telefono,
        "fecha_activa": fecha_busqueda
    })

# ✅ ACTUALIZAR INFORMACIÓN DE VIAJE
@app.post("/prospectos/{prospecto_id}/actualizar-viaje")
async def actualizar_viaje(
    request: Request,
    prospecto_id: int,
    nombre: str = Form(None),  
    apellido: str = Form(None),  
    correo_electronico: str = Form(None),
    telefono: str = Form(...),
    indicativo_telefono: str = Form("57"),
    indicativo_telefono_secundario: str = Form("57"),
    ciudad_origen: str = Form(None),
    destino: str = Form(None),
    fecha_ida: str = Form(None),
    fecha_vuelta: str = Form(None),
    pasajeros_adultos: int = Form(1),
    pasajeros_ninos: int = Form(0),
    pasajeros_infantes: int = Form(0),
    telefono_secundario: str = Form(None),
    fecha_nacimiento: str = Form(None),
    numero_identificacion: str = Form(None),
    db: Session = Depends(database.get_db)
):
    user = await get_current_user(request, db)
    if not user:
        return RedirectResponse(url="/", status_code=303)
    
    try:
        prospecto = db.query(models.Prospecto).filter(models.Prospecto.id == prospecto_id).first()
        if not prospecto:
            return RedirectResponse(
                url=f"/prospectos/{prospecto_id}/seguimiento?error=Prospecto no encontrado", 
                status_code=303
            )
        
        # Verificar permisos
        if (user.tipo_usuario == models.TipoUsuario.AGENTE.value and 
            prospecto.agente_asignado_id != user.id):
            return RedirectResponse(
                url=f"/prospectos/{prospecto_id}/seguimiento?error=No tiene permisos para editar este prospecto", 
                status_code=303
            )
        
        # Convertir fechas
        fecha_ida_date = parsear_fecha(fecha_ida)
        fecha_vuelta_date = parsear_fecha(fecha_vuelta)
        
        # Actualizar información
        prospecto.nombre = nombre
        prospecto.apellido = apellido
        prospecto.correo_electronico = correo_electronico
        prospecto.ciudad_origen = ciudad_origen
        prospecto.destino = destino
        prospecto.telefono = telefono
        prospecto.indicativo_telefono = indicativo_telefono
        prospecto.telefono_secundario = telefono_secundario
        prospecto.indicativo_telefono_secundario = indicativo_telefono_secundario
        prospecto.fecha_ida = fecha_ida_date
        prospecto.fecha_vuelta = fecha_vuelta_date
        prospecto.pasajeros_adultos = pasajeros_adultos
        prospecto.pasajeros_ninos = pasajeros_ninos
        prospecto.pasajeros_infantes = pasajeros_infantes
        prospecto.telefono_secundario = telefono_secundario
        
        # Campos adicionales para clientes ganados
        fecha_nacimiento_date = parsear_fecha(fecha_nacimiento) if fecha_nacimiento else None
        if prospecto.estado == EstadoProspecto.GANADO.value or prospecto.estado == EstadoProspecto.VENTA_CANCELADA.value:
            prospecto.fecha_nacimiento = fecha_nacimiento_date
            prospecto.numero_identificacion = numero_identificacion
        
        # Registrar interacción automática
        interaccion = models.Interaccion(
            prospecto_id=prospecto_id,
            usuario_id=user.id,
            tipo_interaccion="sistema",
            descripcion="Información de viaje actualizada",
            estado_anterior=prospecto.estado,
            estado_nuevo=prospecto.estado
        )
        
        db.add(interaccion)
        db.commit()
        
        return RedirectResponse(
            url=f"/prospectos/{prospecto_id}/seguimiento?success=Información de viaje actualizada correctamente", 
            status_code=303
        )
    
    except Exception as e:
        db.rollback()
        print(f"❌ Error actualizando viaje: {e}")
        return RedirectResponse(
            url=f"/prospectos/{prospecto_id}/seguimiento?error=Error al actualizar información", 
            status_code=303
        )

# Endpoint de salud
@app.get("/health")
async def health_check(db: Session = Depends(database.get_db)):
    try:
        user_count = db.query(models.Usuario).count()
        prospecto_count = db.query(models.Prospecto).count()
        return {
            "status": "healthy",
            "database": "connected", 
            "users_count": user_count,
            "prospectos_count": prospecto_count,
            "active_sessions": len(active_sessions)
        }
    except Exception as e:
        return {
            "status": "unhealthy",
            "error": str(e)
        }

# ========== FILTROS DESDE DASHBOARD ==========
# ✅ NUEVO: FILTRO POR DATOS COMPLETOS/SIN DATOS
@app.get("/prospectos/filtro", response_class=HTMLResponse)
async def prospectos_filtro_dashboard(
    request: Request,
    tipo_filtro: str = Query(...),  # estado, asignacion, destino, ventas, datos, total
    valor_filtro: str = Query(...),  # valor del filtro
    # ✅ AGREGAR PARÁMETROS DE FECHA
    fecha_inicio: str = Query(None),
    fecha_fin: str = Query(None),
    periodo: str = Query("mes"),
    pagina: int = Query(1),
    agente_asignado_id: str = Query(None), # ✅ Nuevo filtro por agente
    db: Session = Depends(database.get_db)
):
    user = await get_current_user(request, db)
    if not user:
        return RedirectResponse(url="/", status_code=303)
    
    # Configurar paginación
    registros_por_pagina = 50
    offset = (pagina - 1) * registros_por_pagina
    
    # ✅ CALCULAR RANGO DE FECHAS (Ya devuelve datetimes con hora min/max)
    fecha_inicio_dt, fecha_fin_dt = calcular_rango_fechas(periodo, fecha_inicio, fecha_fin)
    
    # extraer objetos date para comparaciones que requieran solo fecha (como cotizaciones)
    fecha_inicio_date = fecha_inicio_dt.date()
    fecha_fin_date = fecha_fin_dt.date()
    
    # ✅ CONSTRUIR QUERY SEGÚN TIPO DE FILTRO Y FECHAS
    if tipo_filtro == "estado" and valor_filtro == EstadoProspecto.COTIZADO.value:
        # COTIZADOS: Usar EstadisticaCotizacion (comparar fechas, no datetimes)
        query = db.query(models.Prospecto).join(
            models.EstadisticaCotizacion, 
            models.EstadisticaCotizacion.prospecto_id == models.Prospecto.id
        ).filter(
            models.EstadisticaCotizacion.fecha_cotizacion >= fecha_inicio_date,
            models.EstadisticaCotizacion.fecha_cotizacion <= fecha_fin_date
        )
        if user.tipo_usuario == TipoUsuario.AGENTE.value:
            query = query.filter(models.EstadisticaCotizacion.agente_id == user.id)
            
        titulo_filtro = "Prospectos que han sido cotizados en el periodo"

    elif (tipo_filtro == "estado" and valor_filtro in [
            EstadoProspecto.EN_SEGUIMIENTO.value, 
            EstadoProspecto.GANADO.value, 
            EstadoProspecto.CERRADO_PERDIDO.value
        ]) or (tipo_filtro == "ventas"):
        
        # SEGUIMIENTO, GANADOS, PERDIDOS: Usar HistorialEstado
        target_state = valor_filtro
        if tipo_filtro == "ventas":
            target_state = EstadoProspecto.GANADO.value
            titulo_filtro = "Ventas realizadas en el periodo"
        else:
            titulo_filtro = f"Prospectos {valor_filtro.replace('_', ' ').title()} en el periodo"
            
        query = db.query(models.Prospecto).join(
            models.HistorialEstado,
            models.HistorialEstado.prospecto_id == models.Prospecto.id
        ).filter(
            models.HistorialEstado.estado_nuevo == target_state,
            models.HistorialEstado.fecha_cambio >= fecha_inicio_dt,
            models.HistorialEstado.fecha_cambio <= fecha_fin_dt
        )
        
        if user.tipo_usuario == TipoUsuario.AGENTE.value:
            query = query.filter(models.HistorialEstado.usuario_id == user.id)

    else:
        # OTROS (Nuevos, Total, Asignación, etc.): Usar Fecha de Registro
        query = db.query(models.Prospecto).filter(
            models.Prospecto.fecha_registro >= fecha_inicio_dt,
            models.Prospecto.fecha_registro <= fecha_fin_dt
        )
        
        if user.tipo_usuario == TipoUsuario.AGENTE.value:
            query = query.filter(models.Prospecto.agente_asignado_id == user.id)
    
    # ✅ Filtro por agente explicito (para Admin/Supervisor) - MOVIDO FUERA DEL ELSE
    if agente_asignado_id and agente_asignado_id != "todos":
        try:
            # Si estamos filtrando historia/stats, el filtro de agente es diferente
            if tipo_filtro == "estado" and valor_filtro == EstadoProspecto.COTIZADO.value:
                query = query.filter(models.EstadisticaCotizacion.agente_id == int(agente_asignado_id))
            elif (tipo_filtro == "estado" and valor_filtro in [EstadoProspecto.EN_SEGUIMIENTO.value, EstadoProspecto.GANADO.value, EstadoProspecto.CERRADO_PERDIDO.value]) or (tipo_filtro == "ventas"):
                query = query.filter(models.HistorialEstado.usuario_id == int(agente_asignado_id))
            else:
                query = query.filter(models.Prospecto.agente_asignado_id == int(agente_asignado_id))
        except ValueError:
            pass
            
    # Aplicar filtros específicos adicionales
    # ✅ IMPORTANTE: No aplicar filtro de estado actual para cotizaciones
    # porque ya se manejó arriba usando EstadisticaCotizacion
    if tipo_filtro == "estado" and valor_filtro != EstadoProspecto.COTIZADO.value:
        # Solo queda NUEVO o cualquier otro no manejado arriba
        query = query.filter(models.Prospecto.estado == valor_filtro)
        titulo_filtro = f"Prospectos en estado: {valor_filtro.replace('_', ' ').title()}"
        
    elif tipo_filtro == "asignacion":
        if valor_filtro == "sin_asignar":
            query = query.filter(models.Prospecto.agente_asignado_id == None)
            titulo_filtro = "Prospectos sin asignar"
        elif valor_filtro == "asignados":
            query = query.filter(models.Prospecto.agente_asignado_id != None)
            titulo_filtro = "Prospectos asignados"
    
    elif tipo_filtro == "destino":
        query = query.filter(models.Prospecto.destino.ilike(f"%{valor_filtro}%"))
        titulo_filtro = f"Prospectos con destino: {valor_filtro}"
    
    elif tipo_filtro == "datos":
        if valor_filtro == "con_datos":
            query = query.filter(models.Prospecto.tiene_datos_completos == True)
            titulo_filtro = "Prospectos con datos completos"
        elif valor_filtro == "sin_datos":
            query = query.filter(models.Prospecto.tiene_datos_completos == False)
            titulo_filtro = "Prospectos sin datos (solo teléfono)"
            
    elif tipo_filtro == "total":
        titulo_filtro = "Todos los prospectos registrados"
    
    # Obtener total y prospectos paginados
    total_prospectos = query.count()
    prospectos = query.offset(offset).limit(registros_por_pagina).all()
    
    # Calcular total de páginas
    total_paginas = (total_prospectos + registros_por_pagina - 1) // registros_por_pagina
    
    # Obtener datos para filtros
    agentes = db.query(models.Usuario).filter(
        models.Usuario.tipo_usuario == TipoUsuario.AGENTE.value
    ).all()
    
    medios_ingreso = db.query(models.MedioIngreso).all()
    
    return templates.TemplateResponse("prospectos_filtro.html", {
        "request": request,
        "prospectos": prospectos,
        "current_user": user,
        "agentes": agentes,
        "medios_ingreso": medios_ingreso,
        "titulo_filtro": titulo_filtro,
        "tipo_filtro": tipo_filtro,
        "valor_filtro": valor_filtro,
        "pagina_actual": pagina,
        "total_paginas": total_paginas,
        "total_prospectos": total_prospectos,
        "registros_por_pagina": registros_por_pagina,
        # ✅ PASAR DATOS DE FECHA
        "fecha_inicio_activa": fecha_inicio,
        "fecha_fin_activa": fecha_fin,
        "periodo_activo": periodo,
        "fecha_inicio_formateada": fecha_inicio_dt.strftime("%d/%m/%Y"),
        "fecha_fin_formateada": fecha_fin_dt.strftime("%d/%m/%Y")
    })


# ✅ ENDPOINT PARA AUTOCOMPLETADO DE DESTINOS
@app.get("/api/destinos/sugerencias")
async def sugerencias_destinos(
    q: str = Query("", min_length=2),
    limit: int = Query(10),
    db: Session = Depends(database.get_db)
):
    """Devuelve sugerencias de destinos existentes"""
    if len(q) < 2:
        return JSONResponse(content={"sugerencias": []})
    
    try:
        # Buscar destinos que contengan el texto (case-insensitive)
        destinos = db.query(models.Prospecto.destino).filter(
            models.Prospecto.destino.isnot(None),
            models.Prospecto.destino != '',
            models.Prospecto.destino.ilike(f"%{q}%")
        ).distinct().limit(limit).all()
        
        # Extraer solo los strings
        sugerencias = [destino[0] for destino in destinos if destino[0]]
        
        # Ordenar por relevancia (los que empiezan con la búsqueda primero)
        sugerencias.sort(key=lambda x: 
            0 if x.lower().startswith(q.lower()) else 
            1 if q.lower() in x.lower() else 2
        )
        
        return JSONResponse(content={"sugerencias": sugerencias[:limit]})
        
    except Exception as e:
        print(f"Error en sugerencias_destinos: {e}")
        return JSONResponse(content={"sugerencias": []})

# ✅ ENDPOINT PARA NORMALIZAR DESTINOS EXISTENTES
@app.post("/api/destinos/normalizar")
async def normalizar_destinos(
    destino_original: str = Form(...),
    destino_normalizado: str = Form(...),
    aplicar_a_todos: bool = Form(False),
    db: Session = Depends(database.get_db),
    user: models.Usuario = Depends(get_current_user)
):
    """Normaliza un destino existente"""
    if not user or user.tipo_usuario not in [TipoUsuario.ADMINISTRADOR.value, TipoUsuario.SUPERVISOR.value]:
        raise HTTPException(status_code=403, detail="No tiene permisos")
    
    try:
        if aplicar_a_todos:
            # Actualizar todos los prospectos con este destino
            prospectos = db.query(models.Prospecto).filter(
                models.Prospecto.destino.ilike(f"%{destino_original}%")
            ).all()
            
            for prospecto in prospectos:
                prospecto.destino = destino_normalizado
            
            count = len(prospectos)
            mensaje = f"Se normalizaron {count} prospectos"
        else:
            # Actualizar solo los exactos
            prospectos = db.query(models.Prospecto).filter(
                models.Prospecto.destino == destino_original
            ).all()
            
            for prospecto in prospectos:
                prospecto.destino = destino_normalizado
            
            count = len(prospectos)
            mensaje = f"Se normalizaron {count} prospectos"
        
        db.commit()
        
        # Registrar acción en historial
        if count > 0:
            accion = models.Interaccion(
                prospecto_id=None,  # Acción global
                usuario_id=user.id,
                tipo_interaccion="sistema",
                descripcion=f"Normalización de destinos: '{destino_original}' → '{destino_normalizado}' ({count} registros)",
                estado_anterior=None,
                estado_nuevo=None
            )
            db.add(accion)
            db.commit()
        
        return {"success": True, "message": mensaje, "count": count}
        
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")


# ========== ESTADÍSTICAS AVANZADAS ==========

@app.get("/estadisticas/cotizaciones", response_class=HTMLResponse)
async def estadisticas_cotizaciones(
    request: Request,
    periodo: str = Query("mes"),  # dia, semana, mes, año, personalizado
    fecha_inicio: str = Query(None),
    fecha_fin: str = Query(None),
    agente_id: str = Query(None),
    db: Session = Depends(database.get_db)
):
    user = await get_current_user(request, db)
    if not user:
        return RedirectResponse(url="/", status_code=303)
    
    try:
        # Determinar rango de fechas
        fecha_inicio_dt, fecha_fin_dt = calcular_rango_fechas(periodo, fecha_inicio, fecha_fin)
        fecha_inicio_obj = fecha_inicio_dt.date()
        fecha_fin_obj = fecha_fin_dt.date()
        
        # Construir query base
        query = db.query(
            models.EstadisticaCotizacion,
            models.Usuario.username
        ).join(
            models.Usuario, models.EstadisticaCotizacion.agente_id == models.Usuario.id
        ).filter(
            models.EstadisticaCotizacion.fecha_cotizacion >= fecha_inicio_obj,
            models.EstadisticaCotizacion.fecha_cotizacion <= fecha_fin_obj
        )
        
        # Filtro por agente
        if agente_id and agente_id != "todos":
            query = query.filter(models.EstadisticaCotizacion.agente_id == int(agente_id))
        elif user.tipo_usuario == TipoUsuario.AGENTE.value:
            # Agente solo ve sus propias estadísticas
            query = query.filter(models.EstadisticaCotizacion.agente_id == user.id)
        
        # Agrupar por agente y fecha
        # ✅ CAMBIO: Obtener lista detallada de cotizaciones individualmente
        estadisticas = query.add_columns(
            models.Prospecto.id.label('prospecto_id'),
            models.Prospecto.nombre,
            models.Prospecto.apellido,
            models.Prospecto.telefono
        ).join(
            models.Prospecto, models.EstadisticaCotizacion.prospecto_id == models.Prospecto.id
        ).order_by(
            models.EstadisticaCotizacion.fecha_cotizacion.desc(),
            models.Usuario.username
        ).all()
        
        # Estadísticas resumidas por agente
        resumen_agentes = db.query(
            models.Usuario.id,
            models.Usuario.username,
            func.count(models.EstadisticaCotizacion.id).label('total')
        ).join(
            models.EstadisticaCotizacion, models.EstadisticaCotizacion.agente_id == models.Usuario.id
        ).filter(
            models.EstadisticaCotizacion.fecha_cotizacion >= fecha_inicio_obj,
            models.EstadisticaCotizacion.fecha_cotizacion <= fecha_fin_obj
        )
        
        if user.tipo_usuario == TipoUsuario.AGENTE.value:
            resumen_agentes = resumen_agentes.filter(models.EstadisticaCotizacion.agente_id == user.id)
        
        resumen_agentes = resumen_agentes.group_by(models.Usuario.id, models.Usuario.username).all()
        
        # Obtener lista de agentes para filtro (solo activos)
        agentes = db.query(models.Usuario).filter(
            models.Usuario.tipo_usuario == TipoUsuario.AGENTE.value,
            models.Usuario.activo == 1
        ).all()
        
        return templates.TemplateResponse("estadisticas_cotizaciones.html", {
            "request": request,
            "current_user": user,
            "estadisticas": estadisticas,
            "resumen_agentes": resumen_agentes,
            "agentes": agentes,
            "periodo_activo": periodo,
            "fecha_inicio_activa": fecha_inicio,
            "fecha_fin_activa": fecha_fin,
            "agente_id_activo": agente_id,
            "fecha_inicio_formateada": fecha_inicio_obj.strftime("%d/%m/%Y"),
            "fecha_fin_formateada": fecha_fin_obj.strftime("%d/%m/%Y")
        })
    
    except Exception as e:
        print(f"❌ Error en estadísticas: {e}")
        return RedirectResponse(url="/dashboard?error=Error al cargar estadísticas", status_code=303)


# ========== SISTEMA DE NOTIFICACIONES ==========

def check_inactivity(db: Session):
    """Verifica prospectos nuevos sin gestión por más de 4 horas"""
    limite = datetime.now() - timedelta(hours=4)
    
    # Prospectos nuevos creados antes del limite
    prospectos_inactivos = db.query(models.Prospecto).filter(
        models.Prospecto.estado == EstadoProspecto.NUEVO.value,
        models.Prospecto.fecha_registro <= limite
    ).all()
    
    count = 0
    for p in prospectos_inactivos:
        # Verificar si ya tiene alerta de inactividad reciente (últimas 24h)
        existe_alerta = db.query(models.Notificacion).filter(
            models.Notificacion.prospecto_id == p.id,
            models.Notificacion.tipo == "inactividad",
            models.Notificacion.fecha_creacion >= datetime.now() - timedelta(hours=24)
        ).first()
        
        if not existe_alerta:
            # Buscar supervisor/admin para notificar (o al agente si está asignado, aunque nuevos suelen estar sin asignar)
            # Si tiene agente, notificamos al agente. Si no, a todos los admins/supervisores.
            destinatarios = []
            if p.agente_asignado_id:
                destinatarios.append(p.agente_asignado_id)
            else:
                # Notificar a admins/supervisores
                admins = db.query(models.Usuario).filter(
                    models.Usuario.tipo_usuario.in_([TipoUsuario.ADMINISTRADOR.value, TipoUsuario.SUPERVISOR.value])
                ).all()
                destinatarios = [u.id for u in admins]
            
            for uid in destinatarios:
                notificacion = models.Notificacion(
                    usuario_id=uid,
                    prospecto_id=p.id,
                    tipo="inactividad",
                    mensaje=f"⚠️ Prospecto inactivo > 4h: {p.nombre} {p.apellido or ''}",
                    email_enviado=False
                )
                db.add(notificacion)
                count += 1
    
    db.commit()
    return count

@app.get("/api/notificaciones/check-inactivity")
async def api_check_inactivity(
    db: Session = Depends(database.get_db)
):
    """Endpoint para activar la verificación manual o por cron"""
    try:
        count = check_inactivity(db)
        return {"status": "ok", "alertas_generadas": count}
    except Exception as e:
        return {"status": "error", "message": str(e)}

@app.get("/notificaciones", response_class=HTMLResponse)
async def ver_notificaciones(
    request: Request,
    page: int = Query(1, ge=1),
    limit: int = Query(20, ge=1, le=100),
    filtro_agente_id: str = Query(None),
    filtro_tipo: str = Query(None),  # NUEVO
    filtro_estado: str = Query(None),  # NUEVO
    fecha_inicio: str = Query(None),  # NUEVO
    fecha_fin: str = Query(None),  # NUEVO
    db: Session = Depends(database.get_db)
):
    user = await get_current_user(request, db)
    if not user:
        return RedirectResponse(url="/", status_code=303)
    
    # Trigger inactividad check al cargar (para asegurar alertas frescas)
    check_inactivity(db)
    
    # Query base
    query = db.query(models.Notificacion)
    
    # Filtro por agente
    if user.tipo_usuario == TipoUsuario.AGENTE.value:
        query = query.filter(models.Notificacion.usuario_id == user.id)
    elif filtro_agente_id and filtro_agente_id != "todos":
        query = query.filter(models.Notificacion.usuario_id == int(filtro_agente_id))
    
    # NUEVO: Filtro por tipo
    if filtro_tipo and filtro_tipo != "todos":
        query = query.filter(models.Notificacion.tipo == filtro_tipo)
    
    # NUEVO: Filtro por estado
    if filtro_estado:
        if filtro_estado == "pendientes":
            query = query.filter(models.Notificacion.leida == False)
        elif filtro_estado == "leidas":
            query = query.filter(models.Notificacion.leida == True)
        elif filtro_estado == "vencidas":
            query = query.filter(
                models.Notificacion.leida == False,
                models.Notificacion.fecha_programada.isnot(None),
                models.Notificacion.fecha_programada < datetime.now()
            )
        elif filtro_estado == "proximas":
            query = query.filter(
                models.Notificacion.leida == False,
                models.Notificacion.fecha_programada.isnot(None),
                models.Notificacion.fecha_programada > datetime.now()
            )
    else:
        # Por defecto, solo no leídas
        query = query.filter(models.Notificacion.leida == False)
    
    # NUEVO: Filtro por rango de fechas
    if fecha_inicio:
        try:
            fecha_inicio_dt = datetime.strptime(fecha_inicio, "%d/%m/%Y")
            query = query.filter(models.Notificacion.fecha_creacion >= fecha_inicio_dt)
        except ValueError:
            pass
    
    if fecha_fin:
        try:
            fecha_fin_dt = datetime.strptime(fecha_fin, "%d/%m/%Y")
            fecha_fin_dt = fecha_fin_dt.replace(hour=23, minute=59, second=59)
            query = query.filter(models.Notificacion.fecha_creacion <= fecha_fin_dt)
        except ValueError:
            pass
        
    # Contar total antes de paginar
    total_notificaciones = query.count()
    
    # Calcular paginación
    total_pages = (total_notificaciones + limit - 1) // limit
    offset = (page - 1) * limit
    
    # Obtener notificaciones paginadas
    notificaciones = query.order_by(models.Notificacion.fecha_creacion.desc()).offset(offset).limit(limit).all()
    
    # Calcular tiempos
    for n in notificaciones:
        if n.fecha_programada:
            delta = n.fecha_programada - datetime.now()
            # Formatear tiempo restante
            dias = delta.days
            horas, resto = divmod(delta.seconds, 3600)
            minutos, _ = divmod(resto, 60)
            
            if delta.total_seconds() > 0:
                if dias > 0:
                    n.tiempo_restante_str = f"{dias}d {horas}h"
                else:
                    n.tiempo_restante_str = f"{horas}h {minutos}m"
                n.es_tarde = False
            else:
                n.tiempo_restante_str = "Vencida"
                n.es_tarde = True
        else:
            delta = datetime.now() - n.fecha_creacion
            dias = delta.days
            horas, resto = divmod(delta.seconds, 3600)
            n.tiempo_restante_str = f"Hace {dias}d {horas}h" if dias > 0 else f"Hace {horas}h"
            n.es_tarde = False
            
    agentes = []
    if user.tipo_usuario in [TipoUsuario.ADMINISTRADOR.value, TipoUsuario.SUPERVISOR.value]:
        agentes = db.query(models.Usuario).filter(
            models.Usuario.tipo_usuario == TipoUsuario.AGENTE.value,
            models.Usuario.activo == 1
        ).all()
    
    # NUEVO: Obtener prospectos activos para el modal
    prospectos_activos = []
    if user.tipo_usuario == TipoUsuario.AGENTE.value:
        prospectos_activos = db.query(models.Prospecto).filter(
            models.Prospecto.agente_asignado_id == user.id,
            models.Prospecto.estado.in_(['nuevo', 'en_seguimiento', 'cotizado'])
        ).order_by(models.Prospecto.nombre).limit(50).all()
    else:
        prospectos_activos = db.query(models.Prospecto).filter(
            models.Prospecto.estado.in_(['nuevo', 'en_seguimiento', 'cotizado'])
        ).order_by(models.Prospecto.nombre).limit(100).all()
        
    return templates.TemplateResponse("notificaciones.html", {
        "request": request,
        "current_user": user,
        "notificaciones": notificaciones,
        "agentes": agentes,
        "filtro_agente_id": filtro_agente_id,
        "filtro_tipo": filtro_tipo,
        "filtro_estado": filtro_estado,
        "fecha_inicio": fecha_inicio,
        "fecha_fin": fecha_fin,
        "prospectos_activos": prospectos_activos,
        "page": page,
        "limit": limit,
        "total_pages": total_pages,
        "total_notificaciones": total_notificaciones
    })

@app.post("/notificaciones/{notificacion_id}/leer")
async def marcar_notificacion_leida(
    notificacion_id: int,
    db: Session = Depends(database.get_db),
    request: Request = None 
):
    notif = db.query(models.Notificacion).filter(models.Notificacion.id == notificacion_id).first()
    if notif:
        notif.leida = True
        db.commit()
    
    return RedirectResponse(url="/notificaciones", status_code=303)

# ========== API DE NOTIFICACIONES PUSH ==========

@app.get("/api/notificaciones/pendientes")
async def obtener_notificaciones_pendientes(
    request: Request,
    db: Session = Depends(database.get_db)
):
    """Devuelve notificaciones vencidas y no leídas para el usuario actual"""
    user = await get_current_user(request, db)
    if not user:
        return JSONResponse(content={"notificaciones": []})
    
    ahora = datetime.now()
    
    # Buscar notificaciones programadas que ya vencieron y no han sido leídas
    notificaciones = db.query(models.Notificacion).filter(
        models.Notificacion.usuario_id == user.id,
        models.Notificacion.leida == False,
        models.Notificacion.fecha_programada.isnot(None),
        models.Notificacion.fecha_programada <= ahora
    ).order_by(models.Notificacion.fecha_programada.asc()).all()
    
    # Convertir a JSON
    resultado = []
    for n in notificaciones:
        prospecto_nombre = "N/A"
        if n.prospecto:
            prospecto_nombre = f"{n.prospecto.nombre} {n.prospecto.apellido or ''}".strip()
        
        resultado.append({
            "id": n.id,
            "mensaje": n.mensaje,
            "tipo": n.tipo,
            "prospecto_id": n.prospecto_id,
            "fecha_programada": n.fecha_programada.strftime("%d/%m/%Y %H:%M"),
            "prospecto_nombre": prospecto_nombre
        })
    
    return JSONResponse(content={"notificaciones": resultado})


# ========== ENDPOINTS PARA NOTIFICACIONES MANUALES ==========

@app.get("/api/buscar-prospecto-por-id")
async def buscar_prospecto_por_id(
    id: str = Query(...),
    db: Session = Depends(database.get_db),
    request: Request = None
):
    """Busca un prospecto por ID de Cliente o ID de Cotización"""
    user = await get_current_user(request, db)
    if not user:
        return JSONResponse(content={"success": False, "error": "No autenticado"})
    
    try:
        prospecto = None
        
        # Buscar por ID de Cliente (formato: CL-YYYYMMDD-XXXX)
        if id.startswith('CL-'):
            prospecto = db.query(models.Prospecto).filter(
                models.Prospecto.id_cliente == id
            ).first()
        
        # Buscar por ID de Cotización (formato: COT-YYYYMMDD-XXXX)
        elif id.startswith('COT-'):
            # Buscar en EstadisticaCotizacion
            cotizacion = db.query(models.EstadisticaCotizacion).filter(
                models.EstadisticaCotizacion.id_cotizacion == id
            ).first()
            
            if cotizacion:
                prospecto = cotizacion.prospecto
        
        # Si no tiene prefijo, intentar buscar como número de ID directo
        else:
            try:
                prospecto_id = int(id)
                prospecto = db.query(models.Prospecto).filter(
                    models.Prospecto.id == prospecto_id
                ).first()
            except ValueError:
                pass
        
        # Verificar permisos
        if prospecto:
            if user.tipo_usuario == TipoUsuario.AGENTE.value:
                if prospecto.agente_asignado_id != user.id:
                    return JSONResponse(content={
                        "success": False, 
                        "error": "No tienes permisos para ver este prospecto"
                    })
            
            return JSONResponse(content={
                "success": True,
                "prospecto": {
                    "id": prospecto.id,
                    "id_cliente": prospecto.id_cliente,
                    "nombre": prospecto.nombre,
                    "apellido": prospecto.apellido or "",
                    "destino": prospecto.destino or ""
                }
            })
        else:
            return JSONResponse(content={
                "success": False,
                "error": "Prospecto no encontrado"
            })
    
    except Exception as e:
        print(f"Error buscando prospecto: {e}")
        return JSONResponse(content={
            "success": False,
            "error": "Error en la búsqueda"
        })


@app.post("/notificaciones/crear")
async def crear_notificacion_manual(
    request: Request,
    mensaje: str = Form(...),
    fecha_programada: str = Form(...),
    prospecto_id: int = Form(None),
    db: Session = Depends(database.get_db)
):
    user = await get_current_user(request, db)
    if not user:
        return RedirectResponse(url="/", status_code=303)
    
    try:
        # Parsear fecha
        fecha_prog = datetime.strptime(fecha_programada, "%Y-%m-%dT%H:%M")
        
        # Validar que la fecha sea futura
        if fecha_prog <= datetime.now():
            return RedirectResponse(
                url="/notificaciones?error=La fecha debe ser futura", 
                status_code=303
            )
        
        # Crear notificación
        notificacion = models.Notificacion(
            usuario_id=user.id,
            prospecto_id=prospecto_id if prospecto_id else None,
            tipo="seguimiento",
            mensaje=mensaje,
            fecha_programada=fecha_prog,
            leida=False,
            email_enviado=False
        )
        
        db.add(notificacion)
        
        # Si se asocia a un prospecto, crear interacción automática
        if prospecto_id:
            prospecto = db.query(models.Prospecto).filter(
                models.Prospecto.id == prospecto_id
            ).first()
            
            if prospecto:
                # Formatear fecha para el comentario
                fecha_formateada = fecha_prog.strftime("%d/%m/%Y a las %H:%M")
                
                # Crear interacción/comentario
                interaccion = models.Interaccion(
                    prospecto_id=prospecto_id,
                    usuario_id=user.id,
                    tipo_interaccion="sistema",
                    descripcion=f"📅 Contacto programado para {fecha_formateada}\n{mensaje}",
                    estado_anterior=prospecto.estado,
                    estado_nuevo=prospecto.estado
                )
                
                db.add(interaccion)
        
        db.commit()
        
        return RedirectResponse(
            url="/notificaciones?success=Recordatorio creado correctamente", 
            status_code=303
        )
    
    except ValueError as e:
        return RedirectResponse(
            url="/notificaciones?error=Formato de fecha inválido", 
            status_code=303
        )
    except Exception as e:
        db.rollback()
        print(f"Error creando notificación: {e}")
        return RedirectResponse(
            url="/notificaciones?error=Error al crear recordatorio", 
            status_code=303
        )


# ========== EXPORTACIÓN DE DATOS A EXCEL ==========

def generar_excel_prospectos(prospectos, filename="prospectos_export.xlsx"):
    """
    Genera archivo Excel con lista de prospectos
    """
    try:
        # Preparar datos para DataFrame
        data = []
        for p in prospectos:
            # Obtener última interacción
            ultima_interaccion = ""
            if p.interacciones:
                ultima_int = p.interacciones[0]  # Ya está ordenado por fecha desc
                ultima_interaccion = ultima_int.fecha_creacion.strftime("%d/%m/%Y %H:%M")
            
            data.append({
                'ID Cliente': p.id_cliente or f"CL-{p.id:04d}",
                'Nombre': p.nombre,
                'Apellido': p.apellido,
                'Teléfono': p.telefono or "",
                'Teléfono Secundario': p.telefono_secundario or "",
                'Email': p.correo_electronico or "",
                'Ciudad Origen': p.ciudad_origen or "",
                'Destino': p.destino or "",
                'Fecha Ida': p.fecha_ida.strftime("%d/%m/%Y") if p.fecha_ida else "",
                'Fecha Vuelta': p.fecha_vuelta.strftime("%d/%m/%Y") if p.fecha_vuelta else "",
                'Adultos': p.pasajeros_adultos or 0,
                'Niños': p.pasajeros_ninos or 0,
                'Infantes': p.pasajeros_infantes or 0,
                'Medio Ingreso': p.medio_ingreso.nombre if p.medio_ingreso else "",
                'Agente Asignado': p.agente_asignado.username if p.agente_asignado else "Sin asignar",
                'Estado': p.estado.replace("_", " ").title(),
                'Fecha Registro': p.fecha_registro.strftime("%d/%m/%Y %H:%M"),
                'Última Interacción': ultima_interaccion,
                'Cliente Recurrente': "Sí" if p.cliente_recurrente else "No",
                'Datos Completos': "Sí" if p.tiene_datos_completos else "No",
                'Observaciones': p.observaciones or ""
            })
        
        # Crear DataFrame
        df = pd.DataFrame(data)
        
        # Crear archivo Excel en memoria
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Prospectos', index=False)
            
            # Obtener workbook y worksheet para formato
            workbook = writer.book
            worksheet = writer.sheets['Prospectos']
            
            # Formato de encabezados
            from openpyxl.styles import Font, PatternFill, Alignment
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Auto-ajustar columnas
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Agregar filtros automáticos
            worksheet.auto_filter.ref = worksheet.dimensions
        
        output.seek(0)
        return output
        
    except Exception as e:
        print(f"❌ Error generando Excel de prospectos: {e}")
        import traceback
        traceback.print_exc()
        return None


def generar_excel_estadisticas(stats, periodo, fecha_inicio, fecha_fin, filename="dashboard_export.xlsx"):
    """
    Genera archivo Excel con estadísticas del dashboard
    Múltiples hojas: Resumen, Por Estado, Por Agente, Destinos
    """
    try:
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Hoja 1: Resumen General
            resumen_data = {
                'Métrica': [
                    'Periodo',
                    'Fecha Inicio',
                    'Fecha Fin',
                    '',
                    'Total Prospectos',
                    'Prospectos con Datos Completos',
                    'Prospectos sin Datos',
                    'Clientes Sin Asignar',
                    'Clientes Asignados',
                    'Destinos Únicos',
                    'Ventas Cerradas',
                    '',
                    'Prospectos Nuevos',
                    'En Seguimiento',
                    'Cotizados',
                    'Ganados',
                    'Perdidos',
                    'Ventas Canceladas'
                ],
                'Valor': [
                    periodo.title(),
                    fecha_inicio.strftime("%d/%m/%Y"),
                    fecha_fin.strftime("%d/%m/%Y"),
                    '',
                    stats.get('total_prospectos', 0),
                    stats.get('prospectos_con_datos', 0),
                    stats.get('prospectos_sin_datos', 0),
                    stats.get('clientes_sin_asignar', 0),
                    stats.get('clientes_asignados', 0),
                    stats.get('destinos_count', 0),
                    stats.get('ventas_count', 0),
                    '',
                    stats.get('prospectos_nuevos', 0),
                    stats.get('prospectos_seguimiento', 0),
                    stats.get('prospectos_cotizados', 0),
                    stats.get('prospectos_ganados', 0),
                    stats.get('prospectos_perdidos', 0),
                    stats.get('ventas_canceladas', 0)
                ]
            }
            df_resumen = pd.DataFrame(resumen_data)
            df_resumen.to_excel(writer, sheet_name='Resumen General', index=False)
            
            # Hoja 2: Por Estado
            estados_data = {
                'Estado': ['Nuevo', 'En Seguimiento', 'Cotizado', 'Ganado', 'Perdido', 'Venta Cancelada'],
                'Cantidad': [
                    stats.get('prospectos_nuevos', 0),
                    stats.get('prospectos_seguimiento', 0),
                    stats.get('prospectos_cotizados', 0),
                    stats.get('prospectos_ganados', 0),
                    stats.get('prospectos_perdidos', 0),
                    stats.get('ventas_canceladas', 0)
                ]
            }
            df_estados = pd.DataFrame(estados_data)
            df_estados.to_excel(writer, sheet_name='Por Estado', index=False)
            
            # Hoja 3: Por Agente
            if stats.get('conversion_agentes'):
                agentes_data = []
                for agente in stats['conversion_agentes']:
                    agentes_data.append({
                        'Agente': agente['username'],
                        'Total Prospectos': agente['total_prospectos'],
                        'Cotizados': agente['cotizados'],
                        'Ganados': agente['ganados'],
                        'Tasa Conversión': f"{(agente['ganados'] / agente['total_prospectos'] * 100) if agente['total_prospectos'] > 0 else 0:.1f}%"
                    })
                df_agentes = pd.DataFrame(agentes_data)
                df_agentes.to_excel(writer, sheet_name='Por Agente', index=False)
            
            # Hoja 4: Destinos Populares
            if stats.get('destinos_populares'):
                destinos_data = []
                for destino, count in stats['destinos_populares']:
                    destinos_data.append({
                        'Destino': destino,
                        'Cantidad': count
                    })
                df_destinos = pd.DataFrame(destinos_data)
                df_destinos.to_excel(writer, sheet_name='Destinos Populares', index=False)
            
            # Aplicar formato a todas las hojas
            from openpyxl.styles import Font, PatternFill, Alignment
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            
            for sheet_name in writer.sheets:
                worksheet = writer.sheets[sheet_name]
                
                # Formato de encabezados
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Auto-ajustar columnas
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 40)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
        
        output.seek(0)
        return output
        
    except Exception as e:
        print(f"❌ Error generando Excel de estadísticas: {e}")
        import traceback
        traceback.print_exc()
        return None


def generar_excel_interacciones(interacciones, prospecto, filename="interacciones_export.xlsx"):
    """
    Genera archivo Excel con historial de interacciones
    """
    try:
        data = []
        for interaccion in interacciones:
            data.append({
                'Fecha': interaccion.fecha_creacion.strftime("%d/%m/%Y %H:%M"),
                'Usuario': interaccion.usuario.username if interaccion.usuario else "Sistema",
                'Tipo': interaccion.tipo_interaccion.replace("_", " ").title() if interaccion.tipo_interaccion else "General",
                'Descripción': interaccion.descripcion,
                'Estado Anterior': interaccion.estado_anterior.replace("_", " ").title() if interaccion.estado_anterior else "",
                'Estado Nuevo': interaccion.estado_nuevo.replace("_", " ").title() if interaccion.estado_nuevo else ""
            })
        
        df = pd.DataFrame(data)
        
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Información del prospecto
            info_data = {
                'Campo': ['ID Cliente', 'Nombre', 'Teléfono', 'Email', 'Estado Actual'],
                'Valor': [
                    prospecto.id_cliente or f"CL-{prospecto.id:04d}",
                    f"{prospecto.nombre} {prospecto.apellido}",
                    prospecto.telefono or "",
                    prospecto.correo_electronico or "",
                    prospecto.estado.replace("_", " ").title()
                ]
            }
            df_info = pd.DataFrame(info_data)
            df_info.to_excel(writer, sheet_name='Información', index=False)
            
            # Historial de interacciones
            df.to_excel(writer, sheet_name='Historial', index=False)
            
            # Aplicar formato
            from openpyxl.styles import Font, PatternFill, Alignment
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            
            for sheet_name in writer.sheets:
                worksheet = writer.sheets[sheet_name]
                
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 60)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
        
        output.seek(0)
        return output
        
    except Exception as e:
        print(f"❌ Error generando Excel de interacciones: {e}")
        import traceback
        traceback.print_exc()
        return None


# ========== ENDPOINTS DE EXPORTACIÓN ==========

@app.get("/exportar/prospectos")
async def exportar_prospectos(
    request: Request,
    destino: str = Query(None),
    telefono: str = Query(None),
    medio_ingreso_id: str = Query(None),
    agente_asignado_id: str = Query(None),
    estado: str = Query(None),
    busqueda_global: str = Query(None),
    fecha_inicio: str = Query(None),
    fecha_fin: str = Query(None),
    db: Session = Depends(database.get_db)
):
    """Exporta prospectos a Excel con filtros aplicados"""
    user = await get_current_user(request, db)
    
    if not user:
        return RedirectResponse(url="/", status_code=303)
    
    try:
        # Construir query base
        query = db.query(models.Prospecto).filter(
            models.Prospecto.fecha_eliminacion.is_(None)
        )
        
        # Filtrar por agente si no es admin
        if user.tipo_usuario not in [TipoUsuario.ADMINISTRADOR.value, TipoUsuario.SUPERVISOR.value]:
            query = query.filter(models.Prospecto.agente_asignado_id == user.id)
        
        # Aplicar filtros
        if destino:
            query = query.filter(models.Prospecto.destino.ilike(f"%{destino}%"))
        
        if telefono:
            telefono_normalizado = normalizar_numero(telefono)
            query = query.filter(
                or_(
                    func.replace(func.replace(models.Prospecto.telefono, ' ', ''), '-', '').ilike(f"%{telefono_normalizado}%"),
                    func.replace(func.replace(models.Prospecto.telefono_secundario, ' ', ''), '-', '').ilike(f"%{telefono_normalizado}%")
                )
            )
        
        if medio_ingreso_id and medio_ingreso_id != "todos":
            query = query.filter(models.Prospecto.medio_ingreso_id == int(medio_ingreso_id))
        
        if agente_asignado_id and agente_asignado_id != "todos":
            if agente_asignado_id == "sin_asignar":
                query = query.filter(models.Prospecto.agente_asignado_id.is_(None))
            else:
                query = query.filter(models.Prospecto.agente_asignado_id == int(agente_asignado_id))
        
        if estado and estado != "todos":
            query = query.filter(models.Prospecto.estado == estado)
        
        if busqueda_global:
            busqueda = f"%{busqueda_global}%"
            query = query.filter(
                or_(
                    models.Prospecto.nombre.ilike(busqueda),
                    models.Prospecto.apellido.ilike(busqueda),
                    models.Prospecto.telefono.ilike(busqueda),
                    models.Prospecto.correo_electronico.ilike(busqueda),
                    models.Prospecto.id_cliente.ilike(busqueda)
                )
            )
        
        # Filtros de fecha
        if fecha_inicio:
            try:
                fecha_inicio_obj = datetime.strptime(fecha_inicio, "%d/%m/%Y").date()
                fecha_inicio_dt = datetime.combine(fecha_inicio_obj, datetime.min.time())
                query = query.filter(models.Prospecto.fecha_registro >= fecha_inicio_dt)
            except ValueError:
                pass
        
        if fecha_fin:
            try:
                fecha_fin_obj = datetime.strptime(fecha_fin, "%d/%m/%Y").date()
                fecha_fin_dt = datetime.combine(fecha_fin_obj, datetime.max.time())
                query = query.filter(models.Prospecto.fecha_registro <= fecha_fin_dt)
            except ValueError:
                pass
        
        # Ordenar y obtener resultados
        prospectos = query.order_by(models.Prospecto.fecha_registro.desc()).limit(10000).all()
        
        # Generar Excel
        excel_file = generar_excel_prospectos(prospectos)
        
        if not excel_file:
            raise HTTPException(status_code=500, detail="Error generando archivo Excel")
        
        # Nombre del archivo con timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"prospectos_{timestamp}.xlsx"
        
        return StreamingResponse(
            excel_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        print(f"❌ Error exportando prospectos: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail="Error al exportar datos")


@app.get("/exportar/dashboard")
async def exportar_dashboard(
    request: Request,
    periodo: str = Query("mes"),
    fecha_inicio: str = Query(None),
    fecha_fin: str = Query(None),
    db: Session = Depends(database.get_db),
    user: models.Usuario = Depends(require_admin)
):
    """Exporta estadísticas del dashboard a Excel (solo admins)"""
    try:
        # Calcular rango de fechas
        fecha_inicio_obj, fecha_fin_obj = calcular_rango_fechas(periodo, fecha_inicio, fecha_fin)
        
        # Convertir a datetime
        fecha_inicio_dt = datetime.combine(fecha_inicio_obj, datetime.min.time())
        fecha_fin_dt = datetime.combine(fecha_fin_obj, datetime.max.time())
        
        # Recopilar estadísticas (similar al dashboard)
        stats = {}
        
        # Total de prospectos
        stats['total_prospectos'] = db.query(models.Prospecto).filter(
            models.Prospecto.fecha_registro >= fecha_inicio_dt,
            models.Prospecto.fecha_registro <= fecha_fin_dt
        ).count()
        
        # Prospectos con/sin datos
        stats['prospectos_con_datos'] = db.query(models.Prospecto).filter(
            models.Prospecto.tiene_datos_completos == True,
            models.Prospecto.fecha_registro >= fecha_inicio_dt,
            models.Prospecto.fecha_registro <= fecha_fin_dt
        ).count()
        
        stats['prospectos_sin_datos'] = db.query(models.Prospecto).filter(
            models.Prospecto.tiene_datos_completos == False,
            models.Prospecto.fecha_registro >= fecha_inicio_dt,
            models.Prospecto.fecha_registro <= fecha_fin_dt
        ).count()
        
        # Clientes sin asignar
        stats['clientes_sin_asignar'] = db.query(models.Prospecto).filter(
            models.Prospecto.estado == EstadoProspecto.NUEVO.value,
            models.Prospecto.agente_asignado_id == None,
            models.Prospecto.fecha_registro >= fecha_inicio_dt,
            models.Prospecto.fecha_registro <= fecha_fin_dt
        ).count()
        
        # Clientes asignados
        stats['clientes_asignados'] = db.query(models.Prospecto).filter(
            models.Prospecto.agente_asignado_id != None,
            models.Prospecto.fecha_registro >= fecha_inicio_dt,
            models.Prospecto.fecha_registro <= fecha_fin_dt
        ).count()
        
        # Destinos
        destinos_query = db.query(models.Prospecto.destino).filter(
            models.Prospecto.fecha_registro >= fecha_inicio_dt,
            models.Prospecto.fecha_registro <= fecha_fin_dt,
            models.Prospecto.destino.isnot(None),
            models.Prospecto.destino != ''
        ).distinct().all()
        stats['destinos_count'] = len(destinos_query)
        
        # Ventas
        stats['ventas_count'] = db.query(models.Prospecto).filter(
            models.Prospecto.estado == EstadoProspecto.GANADO.value,
            models.Prospecto.fecha_registro >= fecha_inicio_dt,
            models.Prospecto.fecha_registro <= fecha_fin_dt
        ).count()
        
        # Por estado
        stats['prospectos_nuevos'] = db.query(models.Prospecto).filter(
            models.Prospecto.estado == EstadoProspecto.NUEVO.value,
            models.Prospecto.fecha_registro >= fecha_inicio_dt,
            models.Prospecto.fecha_registro <= fecha_fin_dt
        ).count()
        
        stats['prospectos_seguimiento'] = db.query(models.Prospecto).filter(
            models.Prospecto.estado == EstadoProspecto.EN_SEGUIMIENTO.value,
            models.Prospecto.fecha_registro >= fecha_inicio_dt,
            models.Prospecto.fecha_registro <= fecha_fin_dt
        ).count()
        
        stats['prospectos_cotizados'] = db.query(models.Prospecto).filter(
            models.Prospecto.estado == EstadoProspecto.COTIZADO.value,
            models.Prospecto.fecha_registro >= fecha_inicio_dt,
            models.Prospecto.fecha_registro <= fecha_fin_dt
        ).count()
        
        stats['prospectos_ganados'] = db.query(models.Prospecto).filter(
            models.Prospecto.estado == EstadoProspecto.GANADO.value,
            models.Prospecto.fecha_registro >= fecha_inicio_dt,
            models.Prospecto.fecha_registro <= fecha_fin_dt
        ).count()
        
        stats['prospectos_perdidos'] = db.query(models.Prospecto).filter(
            models.Prospecto.estado == EstadoProspecto.CERRADO_PERDIDO.value,
            models.Prospecto.fecha_registro >= fecha_inicio_dt,
            models.Prospecto.fecha_registro <= fecha_fin_dt
        ).count()
        
        stats['ventas_canceladas'] = db.query(models.Prospecto).filter(
            models.Prospecto.estado == EstadoProspecto.VENTA_CANCELADA.value,
            models.Prospecto.fecha_registro >= fecha_inicio_dt,
            models.Prospecto.fecha_registro <= fecha_fin_dt
        ).count()
        
        # Destinos populares
        stats['destinos_populares'] = db.query(
            models.Prospecto.destino,
            func.count(models.Prospecto.id).label('count')
        ).filter(
            models.Prospecto.fecha_registro >= fecha_inicio_dt,
            models.Prospecto.fecha_registro <= fecha_fin_dt,
            models.Prospecto.destino.isnot(None),
            models.Prospecto.destino != ''
        ).group_by(models.Prospecto.destino).order_by(func.count(models.Prospecto.id).desc()).limit(10).all()
        
        # Conversión por agente
        conversion_agentes = []
        agentes = db.query(models.Usuario).filter(
            models.Usuario.tipo_usuario == TipoUsuario.AGENTE.value,
            models.Usuario.activo == 1
        ).all()
        
        for agente in agentes:
            total_agente = db.query(models.Prospecto).filter(
                models.Prospecto.agente_asignado_id == agente.id,
                models.Prospecto.fecha_registro >= fecha_inicio_dt,
                models.Prospecto.fecha_registro <= fecha_fin_dt
            ).count()
            
            ganados_agente = db.query(models.HistorialEstado).filter(
                models.HistorialEstado.usuario_id == agente.id,
                models.HistorialEstado.estado_nuevo == EstadoProspecto.GANADO.value,
                models.HistorialEstado.fecha_cambio >= fecha_inicio_dt,
                models.HistorialEstado.fecha_cambio <= fecha_fin_dt
            ).count()
            
            cotizados_agente = db.query(models.EstadisticaCotizacion).filter(
                models.EstadisticaCotizacion.agente_id == agente.id,
                models.EstadisticaCotizacion.fecha_cotizacion >= fecha_inicio_obj,
                models.EstadisticaCotizacion.fecha_cotizacion <= fecha_fin_obj
            ).count()
            
            conversion_agentes.append({
                'username': agente.username,
                'total_prospectos': total_agente,
                'cotizados': cotizados_agente,
                'ganados': ganados_agente
            })
        
        stats['conversion_agentes'] = conversion_agentes
        
        # Generar Excel
        excel_file = generar_excel_estadisticas(stats, periodo, fecha_inicio_obj, fecha_fin_obj)
        
        if not excel_file:
            raise HTTPException(status_code=500, detail="Error generando archivo Excel")
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"dashboard_{timestamp}.xlsx"
        
        return StreamingResponse(
            excel_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        print(f"❌ Error exportando dashboard: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail="Error al exportar estadísticas")


@app.get("/exportar/interacciones/{prospecto_id}")
async def exportar_interacciones(
    prospecto_id: int,
    request: Request,
    db: Session = Depends(database.get_db)
):
    """Exporta historial de interacciones de un prospecto"""
    user = await get_current_user(request, db)
    
    if not user:
        return RedirectResponse(url="/", status_code=303)
    
    try:
        # Obtener prospecto
        prospecto = db.query(models.Prospecto).filter(models.Prospecto.id == prospecto_id).first()
        
        if not prospecto:
            raise HTTPException(status_code=404, detail="Prospecto no encontrado")
        
        # Verificar permisos
        if user.tipo_usuario not in [TipoUsuario.ADMINISTRADOR.value, TipoUsuario.SUPERVISOR.value]:
            if prospecto.agente_asignado_id != user.id:
                raise HTTPException(status_code=403, detail="No tiene permiso para ver este prospecto")
        
        # Obtener interacciones
        interacciones = db.query(models.Interaccion).filter(
            models.Interaccion.prospecto_id == prospecto_id
        ).order_by(models.Interaccion.fecha_creacion.desc()).all()
        
        # Generar Excel
        excel_file = generar_excel_interacciones(interacciones, prospecto)
        
        if not excel_file:
            raise HTTPException(status_code=500, detail="Error generando archivo Excel")
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        id_cliente = prospecto.id_cliente or f"CL-{prospecto.id:04d}"
        filename = f"interacciones_{id_cliente}_{timestamp}.xlsx"
        
        return StreamingResponse(
            excel_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Error exportando interacciones: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail="Error al exportar interacciones")


@app.get("/exportar/clientes-ganados")
async def exportar_clientes_ganados(
    request: Request,
    fecha_inicio: str = Query(None),
    fecha_fin: str = Query(None),
    db: Session = Depends(database.get_db),
    user: models.Usuario = Depends(require_admin)
):
    """Exporta clientes ganados con datos completos (solo admins)"""
    try:
        # Query base
        query = db.query(models.Prospecto).filter(
            models.Prospecto.estado == EstadoProspecto.GANADO.value,
            models.Prospecto.fecha_eliminacion.is_(None)
        )
        
        # Filtros de fecha
        if fecha_inicio:
            try:
                fecha_inicio_obj = datetime.strptime(fecha_inicio, "%d/%m/%Y").date()
                fecha_inicio_dt = datetime.combine(fecha_inicio_obj, datetime.min.time())
                query = query.filter(models.Prospecto.fecha_compra >= fecha_inicio_dt)
            except ValueError:
                pass
        
        if fecha_fin:
            try:
                fecha_fin_obj = datetime.strptime(fecha_fin, "%d/%m/%Y").date()
                fecha_fin_dt = datetime.combine(fecha_fin_obj, datetime.max.time())
                query = query.filter(models.Prospecto.fecha_compra <= fecha_fin_dt)
            except ValueError:
                pass
        
        # Obtener clientes ganados
        clientes = query.order_by(models.Prospecto.fecha_compra.desc()).all()
        
        # Generar Excel
        excel_file = generar_excel_prospectos(clientes)
        
        if not excel_file:
            raise HTTPException(status_code=500, detail="Error generando archivo Excel")
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"clientes_ganados_{timestamp}.xlsx"
        
        return StreamingResponse(
            excel_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        print(f"❌ Error exportando clientes ganados: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail="Error al exportar clientes ganados")


def generar_excel_usuarios(usuarios, filename="usuarios_export.xlsx"):
    """
    Genera archivo Excel con lista de usuarios del sistema
    """
    try:
        data = []
        for u in usuarios:
            # Contar prospectos asignados
            prospectos_count = len([p for p in u.prospectos if p.fecha_eliminacion is None]) if hasattr(u, 'prospectos') else 0
            
            data.append({
                'ID': u.id,
                'Username': u.username,
                'Email': u.email,
                'Tipo Usuario': u.tipo_usuario.title(),
                'Estado': 'Activo' if u.activo else 'Inactivo',
                'Fecha Creación': u.fecha_creacion.strftime("%d/%m/%Y %H:%M"),
                'Prospectos Asignados': prospectos_count
            })
        
        df = pd.DataFrame(data)
        
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Usuarios', index=False)
            
            workbook = writer.book
            worksheet = writer.sheets['Usuarios']
            
            from openpyxl.styles import Font, PatternFill, Alignment
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 40)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            worksheet.auto_filter.ref = worksheet.dimensions
        
        output.seek(0)
        return output
        
    except Exception as e:
        print(f"❌ Error generando Excel de usuarios: {e}")
        import traceback
        traceback.print_exc()
        return None


@app.get("/exportar/usuarios")
async def exportar_usuarios(
    request: Request,
    db: Session = Depends(database.get_db),
    user: models.Usuario = Depends(require_admin)
):
    """Exporta lista de usuarios a Excel (solo admins)"""
    try:
        # Obtener todos los usuarios
        usuarios = db.query(models.Usuario).order_by(models.Usuario.fecha_creacion.desc()).all()
        
        # Generar Excel
        excel_file = generar_excel_usuarios(usuarios)
        
        if not excel_file:
            raise HTTPException(status_code=500, detail="Error generando archivo Excel")
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"usuarios_{timestamp}.xlsx"
        
        return StreamingResponse(
            excel_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        print(f"❌ Error exportando usuarios: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail="Error al exportar usuarios")


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000, log_level="debug")